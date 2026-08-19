# -*- coding: utf-8 -*-
"""Regressao ponta a ponta: preenche a Kaori na ficha GERADA, manda o
LibreOffice recalcular, e compara com a ficha-exemplo-kaori.docx.

Isso e diferente do conferir-kaori.py, que confere o CATALOGO. Aqui quem esta
sendo julgada e a planilha: as formulas dela, as referencias de celula dela.
Foi assim que apareceu o bug de a Vida somar Destreza no lugar de Constituicao.
"""
import json, os, shutil, subprocess, sys, tempfile, csv
from openpyxl import load_workbook

ARQ = "ficha/ficha-projeto-m.xlsx"
if not os.path.exists(ARQ):
    print("gere a ficha antes:  python3 ficha/monta.py"); sys.exit(1)

KAORI = {"Força": 3, "Constituição": 2, "Destreza": 2, "Inteligência": 1, "Essência": 1}
CAMINHO, NIVEL = "Bastião", 2
# a chave e o nome no INDICE que a propria ficha publica na DADOS
ESPERADO = {"vida_max": 23, "energia_max": 8, "integridade_max": 28, "defesa": 13,
            "maestria": 1, "cd de feitiço": 13}

d = tempfile.mkdtemp(prefix="kaori-")
copia = os.path.join(d, "kaori.xlsx")
shutil.copy(ARQ, copia)
wb = load_workbook(copia)
ws = wb["FICHA"]

# acha as celulas pelo rotulo impresso, e nao por coordenada decorada
def col_do_atributo():
    m = {}
    for linha in ws.iter_rows():
        for c in linha:
            if isinstance(c.value, str) and c.value in KAORI:
                m[c.value] = (c.column, c.row - 4)      # o numero fica 4 linhas acima
    return m

def acha_rotulo(txt):
    for linha in ws.iter_rows():
        for c in linha:
            if isinstance(c.value, str) and c.value.upper() == txt.upper():
                return c.column, c.row
    return None, None

# o indice que a ficha publica: nada de coordenada decorada aqui
IDX = {}
dd = wb["DADOS"]
for r in range(5, 200):
    k, v = dd.cell(row=r, column=53).value, dd.cell(row=r, column=54).value
    if k and v:
        IDX[k] = v
if not IDX:
    print("a ficha nao publicou o indice de celulas; regere com monta.py"); sys.exit(1)

for nome, v in KAORI.items():
    ws[IDX["atr_" + nome]] = v
ws[IDX["caminho"]] = CAMINHO
ws[IDX["nivel"]]   = NIVEL
# o LibreOffice exporta em csv SO a primeira aba, e a primeira agora e a
# CARTEIRA. Na copia, a FICHA vai para a frente -- o arquivo real nao muda.
wb.move_sheet("FICHA", -wb.sheetnames.index("FICHA"))
wb.save(copia)

subprocess.run(["libreoffice", "--headless", "--convert-to",
                "csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true",
                "--outdir", d, copia],
               capture_output=True, timeout=240)
csvf = os.path.join(d, "kaori.csv")
if not os.path.exists(csvf):
    print("o LibreOffice nao converteu; sem ele esta checagem NAO roda")
    sys.exit(1)

linhas = list(csv.reader(open(csvf, encoding="utf-8")))
def le(coord):
    """le o valor recalculado da celula, pelo endereco que o indice deu"""
    col = "".join(ch for ch in coord if ch.isalpha())
    lin = int("".join(ch for ch in coord if ch.isdigit()))
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    if lin - 1 < len(linhas) and n - 1 < len(linhas[lin - 1]):
        return linhas[lin - 1][n - 1].strip()
    return "<fora do csv>"

print(f"Kaori · {CAMINHO} nível {NIVEL} · " +
      " ".join(f"{k[:3]}{v}" for k, v in KAORI.items()))
print(f"\n{"campo":16} {"a ficha calcula":>16} {"manual p.41":>13}")
falhas = 0
for campo, esp in ESPERADO.items():
    lido = le(IDX[campo])
    ok = lido.replace(".0", "") == str(esp)
    falhas += not ok
    print(f"  {campo:16} {lido:>16} {esp:>13}   {'BATE' if ok else 'NÃO BATE'}")

shutil.rmtree(d, ignore_errors=True)
print(f"\n{'A FICHA REPRODUZ A KAORI' if not falhas else f'{falhas} NÚMERO(S) ERRADO(S)'}")
sys.exit(1 if falhas else 0)
