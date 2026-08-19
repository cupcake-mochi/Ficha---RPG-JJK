# -*- coding: utf-8 -*-
"""Helpers dos mockups. Mesma paleta e mesmas fontes ja decididas (A5 e C4):
o que muda entre os mockups e o DESENHO, nunca a decisao."""
import json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEC = json.load(open(os.path.join(RAIZ, "decisoes-ficha.json"), encoding="utf-8"))
G = {g["nome"]: g["hex"] for g in DEC["A5_acento"]["degraus"]}
OSSO, AMBAR, VERMELHO = G["osso"], G["ambar"], G["vermelho"]
CORPO  = DEC["C4_fontes"]["corpo"]["fonte"]
TITULO = DEC["C4_fontes"]["titulo"]["fonte"]
MARCA  = DEC["C4_fontes"]["marca"]["fonte"]

FUNDO, PAINEL, ALTO   = "120F1D", "211C35", "30294D"
LINHA, BLOCO, FRACO   = "493F54", "756588", "998BA9"
TEXTO                 = "F4F1F7"
TINTA                 = "0A0810"          # mais escuro que o fundo: a "tinta"

def fill(c): return PatternFill("solid", start_color=c, end_color=c)

def pinta(ws, c1, r1, c2, r2, cor):
    f = fill(cor)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = f

def junta(ws, c1, r1, c2, r2):
    if (c1, r1) != (c2, r2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def txt(ws, c, r, v, nome=None, pt=11, cor=TEXTO, neg=False,
        al="center", va="center", rot=0, fundo=None, ate=None):
    if ate:
        junta(ws, c, r, ate[0], ate[1])
    cel = ws.cell(row=r, column=c, value=v)
    cel.font = Font(name=nome or CORPO, size=pt, color=cor, bold=neg)
    cel.alignment = Alignment(horizontal=al, vertical=va, textRotation=rot, wrap_text=False)
    if fundo:
        cel.fill = fill(fundo)
    return cel

def regua(ws, c1, r, c2, cor=LINHA, grossa=False):
    """uma linha horizontal, sem caixa em volta: e o que separa desenho de formulario"""
    s = Side(style="medium" if grossa else "thin", color=cor)
    for c in range(c1, c2 + 1):
        cel = ws.cell(row=r, column=c)
        cel.border = Border(top=s)

def barra(n, de, largura=20):
    """medidor em blocos. Funciona em qualquer lugar, inclusive no celular."""
    cheio = 0 if de <= 0 else max(0, min(largura, round(largura * n / de)))
    return "█" * cheio + "─" * (largura - cheio)

def cor_do_estado(n, de):
    if de <= 0: return OSSO
    f = n / de
    return VERMELHO if f <= 0.25 else (AMBAR if f <= 0.5 else OSSO)

def base(wb, nome, cols=46, linhas=70, larg=4.0, alt=15):
    from openpyxl.utils import get_column_letter as L
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    for c in range(1, cols + 1):
        ws.column_dimensions[L(c)].width = larg
    for r in range(1, linhas + 1):
        ws.row_dimensions[r].height = alt
    pinta(ws, 1, 1, cols, linhas, FUNDO)
    return ws
