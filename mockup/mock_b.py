# -*- coding: utf-8 -*-
"""MOCKUP B · a CARTEIRA: a primeira pagina, o documento da pessoa.
Foto, nome, numero de registro, e os carimbos. A parte que nao tem numero
derivado nenhum e existe so para a ficha ser um objeto do mundo."""
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as Img
from openpyxl.utils import get_column_letter as L
from comum import *

ARTE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "arte")
MINCHO, MONO = "Shippori Mincho", "Courier New"

C = 46
wb = Workbook(); wb.remove(wb.active)
ws = base(wb, "CARTEIRA", C, 46)

def arte(nome, col, lin, larg_px, alt_px=None):
    im = Img(os.path.join(ARTE, nome))
    p = im.height / im.width
    im.width = larg_px
    im.height = alt_px if alt_px else int(larg_px * p)
    a = ws.cell(row=lin, column=col).coordinate
    im.anchor = a
    ws.add_image(im)
    return im

# ---------------------------------------------------------------- o papel
pinta(ws, 1, 1, C, 46, TINTA)
arte("textura.png", 1, 1, 1300)

# faixa de cabecalho, mais clara, com o nome da instituicao
pinta(ws, 1, 1, C, 4, "17131F")
txt(ws, 3, 2, "呪術廻戦", nome="Yuji Syuku", pt=22, cor=BLOCO, al="left", ate=(9, 3))
txt(ws, 11, 2, "GUILDA DE FEITICEIROS", nome="Oswald", pt=13, cor=OSSO, al="left", ate=(30, 2))
txt(ws, 11, 3, "carteira de registro · documento oficial", nome=CORPO, pt=9,
    cor=FRACO, al="left", ate=(30, 3))
txt(ws, 36, 2, "Nº M-0104-0027", nome=MONO, pt=11, cor=BLOCO, al="right", ate=(C, 2))
txt(ws, 36, 3, "emitida 18.08.2026", nome=MONO, pt=9, cor=FRACO, al="right", ate=(C, 3))
arte("pincelada-roxa.png", 1, 5, 1240, 14)

# ---------------------------------------------------------------- a foto
arte("moldura-foto.png", 4, 8, 210)

# ---------------------------------------------------------------- o nome
txt(ws, 14, 8, "PORTADOR", nome="Oswald", pt=9, cor=FRACO, al="left", ate=(24, 8))
txt(ws, 14, 9, "Kaori", nome=MINCHO, pt=32, cor=OSSO, al="left", ate=(34, 12))
txt(ws, 14, 13, "カオリ", nome="Shippori Mincho", pt=13, cor=FRACO, al="left", ate=(24, 13))
arte("pincelada.png", 14, 15, 560, 11)

campos = [("CAMINHO", "Bastião"), ("TRILHA", "Escudo"), ("ORIGEM", "Latente"),
          ("NÍVEL", "2"), ("GRAU DA FERRAMENTA", "4"), ("MESA DE ORIGEM", "Rhodes")]
for i, (rot, v) in enumerate(campos):
    c1 = 14 + (i % 3) * 11
    rr = 17 + (i // 3) * 4
    txt(ws, c1, rr, rot, nome="Oswald", pt=8, cor=FRACO, al="left", ate=(c1 + 9, rr))
    txt(ws, c1, rr + 1, v, nome=MINCHO, pt=15, cor=TEXTO, al="left", ate=(c1 + 9, rr + 1))
    regua(ws, c1, rr + 2, c1 + 9, LINHA)

# ---------------------------------------------------------------- o selo
arte("selo-封.png", 37, 24, 120)
txt(ws, 37, 31, "SELO REGISTRADO", nome="Oswald", pt=8, cor=FRACO, ate=(C, 32))

# ---------------------------------------------------------------- rodape
txt(ws, 4, 26, "TÉCNICA DECLARADA", nome="Oswald", pt=9, cor=FRACO, al="left", ate=(20, 26))
txt(ws, 4, 27, "Muralha de Sal", nome=MINCHO, pt=17, cor=OSSO, al="left", ate=(24, 28))
txt(ws, 4, 30, "a energia endurece o que ela toca; o que endurece, protege",
    nome=CORPO, pt=10, cor=FRACO, al="left", ate=(30, 30))
arte("respingo.png", 30, 26, 110)

regua(ws, 3, 36, C, LINHA)
txt(ws, 3, 37, "esta carteira acompanha o portador entre mesas. a ficha completa "
               "está nas páginas seguintes.", nome=CORPO, pt=9, cor=FRACO,
    al="left", ate=(30, 37))
txt(ws, 34, 37, "M-0104-0027-BAS-02", nome=MONO, pt=9, cor=LINHA, al="right", ate=(C, 37))

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_area = f"A1:{L(C)}40"
wb.save("mock-b-carteira.xlsx")
print("mock B escrito")
