# -*- coding: utf-8 -*-
"""MOCKUP A · 'O Registro'.
A ficha e um documento da Guilda: cabecalho carimbado, secoes numeradas,
lombada de kanji na lateral, regua no lugar de caixa, e o Selo como carimbo.
Preenchida com a Kaori para o desenho ser julgado com dado dentro."""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as L
from comum import *

C = 46
wb = Workbook(); wb.remove(wb.active)
ws = base(wb, "REGISTRO", C, 74)
for r in (1,2,3,4,5,6,7): ws.row_dimensions[r].height = 17

# ---------------------------------------------------------------- lombada
# a coluna 1-2 e uma faixa escura que corre a ficha inteira, com kanji girado.
pinta(ws, 1, 1, 2, 74, TINTA)
txt(ws, 1, 8, "呪 術 廻 戦", nome=MARCA, pt=13, cor=LINHA, rot=90, ate=(2, 30))
txt(ws, 1, 34, "PROJETO M", nome=TITULO, pt=11, cor=LINHA, rot=90, ate=(2, 52))

# ---------------------------------------------------------------- cabecalho
pinta(ws, 4, 2, C, 7, TINTA)
txt(ws, 4, 2, "GUILDA · REGISTRO DE FEITICEIRO", nome=TITULO, pt=9, cor=BLOCO,
    al="left", ate=(24, 2))
txt(ws, 4, 3, "KAORI", nome=TITULO, pt=30, cor=OSSO, al="left", ate=(24, 6))
txt(ws, 4, 7, "bastião · trilha do escudo · latente", nome=CORPO, pt=9, cor=FRACO,
    al="left", ate=(24, 7))
# o carimbo de registro, a direita: numero de controle de documento
pinta(ws, 34, 2, C, 7, "17131F")
txt(ws, 34, 3, "REGISTRO", nome=CORPO, pt=8, cor=FRACO, ate=(C, 3))
txt(ws, 34, 4, "M-0104-KAORI", nome=TITULO, pt=13, cor=BLOCO, ate=(C, 5))
txt(ws, 34, 6, "catálogo v0.104 · em dia", nome=CORPO, pt=8, cor=FRACO, ate=(C, 6))
regua(ws, 4, 8, C, BLOCO, grossa=True)

def secao(r, num, nome):
    """o numero da secao entra GRANDE e apagado: e ornamento que tambem informa"""
    txt(ws, 4, r, num, nome=TITULO, pt=22, cor=ALTO, al="left", ate=(6, r + 1))
    txt(ws, 7, r, nome, nome=TITULO, pt=12, cor=BLOCO, al="left", ate=(20, r))
    return r + 2

# ---------------------------------------------------------------- 01 corpo
r = secao(10, "01", "O CORPO")
KA = {"FOR": 3, "DES": 2, "CON": 2, "INT": 1, "ESS": 1}
for i, (a, v) in enumerate(KA.items()):
    c1 = 4 + i * 7
    txt(ws, c1, r, str(v), nome=TITULO, pt=26, cor=OSSO, al="left", ate=(c1 + 5, r + 1))
    txt(ws, c1, r + 2, a, nome=TITULO, pt=9, cor=FRACO, al="left", ate=(c1 + 5, r + 2))
    regua(ws, c1, r + 3, c1 + 4, LINHA)
r += 5

# as tres reservas: numero grande + medidor em blocos, sem caixa nenhuma
for nome_r, at, mx in [("VIDA", 23, 23), ("ENERGIA", 5, 8), ("INTEGRIDADE", 7, 28)]:
    txt(ws, 4, r, nome_r, nome=TITULO, pt=9, cor=FRACO, al="left", ate=(11, r))
    txt(ws, 4, r + 1, str(at), nome=TITULO, pt=24, cor=cor_do_estado(at, mx),
        al="left", ate=(8, r + 2))
    txt(ws, 9, r + 1, f"/ {mx}", nome=CORPO, pt=10, cor=FRACO, al="left", ate=(11, r + 2))
    txt(ws, 13, r + 1, barra(at, mx, 26), nome=CORPO, pt=11,
        cor=cor_do_estado(at, mx), al="left", ate=(30, r + 2))
    if nome_r == "INTEGRIDADE":
        txt(ws, 32, r + 1, "estágio 3 · desvantagem em ataques e TRs",
            nome=CORPO, pt=9, cor=VERMELHO, al="left", ate=(C, r + 2))
    r += 4

# ---------------------------------------------------------------- 02 numeros
r = secao(r, "02", "O QUE CAI SOZINHO")
NUM = [("DEFESA", "13"), ("INICIATIVA", "d20+2"), ("CD", "13"), ("CONJURAÇÃO", "d20+3"),
       ("CORPO A CORPO", "d20+3"), ("À DISTÂNCIA", "d20+2"), ("MAESTRIA", "1"),
       ("DESLOCAMENTO", "9 m")]
for i, (rot, v) in enumerate(NUM):
    c1 = 4 + (i % 4) * 11
    rr = r + (i // 4) * 4
    txt(ws, c1, rr, rot, nome=TITULO, pt=8, cor=FRACO, al="left", ate=(c1 + 9, rr))
    txt(ws, c1, rr + 1, v, nome=TITULO, pt=17, cor=OSSO, al="left", ate=(c1 + 9, rr + 1))
    regua(ws, c1, rr + 2, c1 + 9, LINHA)
r += 8

# ---------------------------------------------------------------- 03 o selo
r = secao(r, "03", "A TÉCNICA")
# o Selo desenhado como carimbo: quadrado, borda dupla, kanji dentro
pinta(ws, 4, r, 12, r + 6, TINTA)
from openpyxl.styles import Border, Side
s = Side(style="medium", color=VERMELHO)
for cc in range(4, 13):
    ws.cell(row=r, column=cc).border = Border(top=s)
    ws.cell(row=r + 6, column=cc).border = Border(bottom=s)
for rr in range(r, r + 7):
    ws.cell(row=rr, column=4).border = Border(left=s,
        top=s if rr == r else None, bottom=s if rr == r + 6 else None)
    ws.cell(row=rr, column=12).border = Border(right=s,
        top=s if rr == r else None, bottom=s if rr == r + 6 else None)
txt(ws, 5, r + 1, "封", nome=MARCA, pt=26, cor=VERMELHO, ate=(11, r + 4))
txt(ws, 5, r + 5, "SELO", nome=TITULO, pt=9, cor=VERMELHO, ate=(11, r + 5))

txt(ws, 14, r, "FAMÍLIAS LIVRES", nome=TITULO, pt=8, cor=FRACO, al="left", ate=(24, r))
txt(ws, 14, r + 1, "Controle · Castigo", nome=CORPO, pt=11, cor=OSSO, al="left", ate=(24, r + 1))
txt(ws, 14, r + 3, "FECHADAS", nome=TITULO, pt=8, cor=FRACO, al="left", ate=(24, r + 3))
txt(ws, 14, r + 4, "Área · Auxiliares · Amparo", nome=CORPO, pt=11, cor=FRACO,
    al="left", ate=(24, r + 4))
txt(ws, 27, r, "CLASSE MÁXIMA", nome=TITULO, pt=8, cor=FRACO, al="left", ate=(33, r))
txt(ws, 27, r + 1, "1", nome=TITULO, pt=26, cor=OSSO, al="left", ate=(33, r + 3))
txt(ws, 35, r, "ESPAÇOS", nome=TITULO, pt=8, cor=FRACO, al="left", ate=(C, r))
txt(ws, 35, r + 1, "3", nome=TITULO, pt=26, cor=OSSO, al="left", ate=(C, r + 3))
r += 8

# um feitico, do jeito que ele aparece na mesa
regua(ws, 4, r, C, LINHA)
r += 1
for nome_f, cl, dano, pe, alc in [("Marca do Carrasco", 1, "3d8", 3, "18 m"),
                                   ("Domo de Gelo", 1, "2d8", 3, "raio 3 m")]:
    txt(ws, 4, r, nome_f, nome=TITULO, pt=13, cor=TEXTO, al="left", ate=(18, r))
    txt(ws, 19, r, f"CL {cl}", nome=CORPO, pt=9, cor=FRACO, ate=(22, r))
    txt(ws, 24, r, dano, nome=TITULO, pt=13, cor=OSSO, ate=(28, r))
    txt(ws, 30, r, f"{pe} PE", nome=CORPO, pt=9, cor=FRACO, ate=(33, r))
    txt(ws, 35, r, alc, nome=CORPO, pt=9, cor=FRACO, al="left", ate=(C, r))
    r += 2

wb.save("mock-a-registro.xlsx")
print("mock A escrito")
