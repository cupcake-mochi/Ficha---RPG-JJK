# -*- coding: utf-8 -*-
"""Monta a Ficha (PROJETO M) 0.1 em .xlsx, a partir do layout.json.

    python3 ficha-v01/extrair.py    # o desenho do .xlsx do Mizuki -> layout.json
    python3 ficha-v01/monta.py      # layout.json -> ficha-projeto-m-0.1.xlsx
    python3 comparar-ficha-01.py    # prova que os dois batem

Por que o desenho mora num JSON e nao em codigo: esta ficha nasceu editada a
mao no Google Sheets -- 8 mil celulas, 317 mesclagens, 105 estilos. Escrever
isso como chamada de funcao seria transcrever um dump com sintaxe de Python, e
qualquer edicao futura dela vai continuar acontecendo no Sheets, nao aqui. O
JSON e o desenho; este arquivo e a maquina que o toca.

Onde a ficha vive e o Google Sheets, e isso decide duas coisas:
  · o IFS das 16 celulas fica CRU. No Excel ele daria #NAME?, e esta aceito.
  · o SPARKLINE das 3 celulas continua. Ele nao existe no Excel.
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.drawing.image import Image as Img

AQUI = os.path.dirname(os.path.abspath(__file__))
LAYOUT = json.load(open(os.path.join(AQUI, "layout.json"), encoding="utf-8"))

# a paleta e a fonte de corpo saem do estilo.py, que e o dono delas -- e ele
# ganhou as quatro cores desta versao na v0.1 (decisao do Mizuki: uma paleta so)
sys.path.insert(0, os.path.join(os.path.dirname(AQUI), "ficha"))
from estilo import CORPO, PT_VALOR, TEXTO

wb = Workbook()
wb.remove(wb.active)
# a fonte padrao do DOCUMENTO, e nao so das celulas com texto: sem isto o
# Excel/Sheets poe a dele em toda celula vazia que leve preenchimento, e foi
# assim que 3165 celulas voltaram em Arial 10 do Sheets.
wb._fonts[0] = Font(name=CORPO, size=PT_VALOR, color=TEXTO)

def monta_fonte(v):
    if not v:
        return None
    nome, tam, cor, neg, ital = v
    return Font(name=nome, size=tam, color=cor, bold=neg, italic=ital)

def monta_borda(v):
    if not v:
        return None
    return Border(**{lado: Side(style=est, color=cor)
                     for lado, (est, cor) in v.items()})

ESTILOS = []
for e in LAYOUT["estilos"]:
    fonte, fundo, bordas, alinha, fmt = e
    ESTILOS.append({
        "font": monta_fonte(fonte),
        "fill": PatternFill("solid", start_color=fundo, end_color=fundo) if fundo else None,
        "border": monta_borda(bordas),
        "alignment": (Alignment(horizontal=alinha[0], vertical=alinha[1],
                                wrap_text=alinha[2], textRotation=alinha[3])
                      if alinha else None),
        "number_format": fmt,
    })

for a in LAYOUT["abas"]:
    ws = wb.create_sheet(a["nome"])
    ws.sheet_state = a["estado"]
    ws.sheet_view.showGridLines = a["grade"]
    if a.get("altura_padrao"):
        ws.sheet_format.defaultRowHeight = a["altura_padrao"]
        ws.sheet_format.customHeight = True

    for cmin, cmax, larg in a["colunas_larg"]:
        for c in range(cmin, cmax + 1):
            ws.column_dimensions[L(c)].width = larg
    for lin, alt in a["linhas_alt"]:
        ws.row_dimensions[lin].height = alt

    # o valor ANTES da mesclagem: mesclar apaga o estilo de tudo que nao e o
    # canto, e ai o bloco abre branco
    for coord, valor, ie in a["celulas"]:
        cel = ws[coord]
        if valor is not None:
            cel.value = valor
        if ie is not None:
            st = ESTILOS[ie]
            if st["font"]:          cel.font = st["font"]
            if st["fill"]:          cel.fill = st["fill"]
            if st["border"]:        cel.border = st["border"]
            if st["alignment"]:     cel.alignment = st["alignment"]
            if st["number_format"]: cel.number_format = st["number_format"]

    for faixa in a["mescladas"]:
        ws.merge_cells(faixa)

    for m in a["menus"]:
        v = DataValidation(type=m["tipo"], formula1=m["formula"],
                           allow_blank=m["vazio_ok"],
                           showDropDown=not m["mostra_seta"])
        ws.add_data_validation(v)
        for parte in m["onde"].split():
            v.add(parte)

    for cf in a["condicional"]:
        dxf = DifferentialStyle(
            font=Font(color=cf["cor_texto"]) if cf["cor_texto"] else None,
            fill=PatternFill(bgColor=cf["fundo"]) if cf["fundo"] else None)
        ws.conditional_formatting.add(
            cf["onde"], Rule(type=cf["tipo"], dxf=dxf, formula=cf["formula"]))

    for im in a["imagens"]:
        caminho = os.path.join(AQUI, "arte", im["arquivo"])
        if not os.path.exists(caminho):
            print(f"  [aviso] falta a arte {im['arquivo']}")
            continue
        img = Img(caminho)
        img.width, img.height = im["larg"], im["alt"]
        img.anchor = ws.cell(row=im["lin"], column=im["col"]).coordinate
        ws.add_image(img)

saida = os.path.join(AQUI, "ficha-projeto-m-0.1.xlsx")
wb.save(saida)
print(f"ficha escrita: {saida}")
print(f"abas: {wb.sheetnames}")
