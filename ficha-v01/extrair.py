# -*- coding: utf-8 -*-
"""Le a Ficha (PROJETO M) 0.1 e escreve o layout.json que o monta.py replica.

Ele NAO copia bytes: ele extrai o DESENHO e limpa o ruido que o Google Sheets
deixou na ida e volta. As tres limpezas, todas decididas pelo Mizuki:

  1. as celulas em Arial 10 preto -- 3165 delas, TODAS vazias. E o estilo de
     fabrica do Sheets em celula que o gerador nao pintou.
  2. a largura 3,63 volta a ser 4,0. O gerador escreve 4,0, o Sheets converte
     para pixel e devolve 3,63: e a mesma coluna.
  3. a condicional verde B7E1CD em D114 -- o "nao esta vazio" de fabrica do
     Google, numa linha sem conteudo.

O que ele PRESERVA: valor e formula de cada celula, as cinco fontes com
tamanho e cor, os preenchimentos, as bordas, o alinhamento, as mesclagens, as
alturas de linha, os menus suspensos, a formatacao condicional de estado da
decisao A5, as imagens e a ordem das abas.
"""
import json, os, re, shutil, sys, zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as L

AQUI = os.path.dirname(os.path.abspath(__file__))
ORIG = sys.argv[1] if len(sys.argv) > 1 else os.path.join(AQUI, "original.xlsx")
SAIDA = os.path.join(AQUI, "layout.json")
ARTE = os.path.join(AQUI, "arte")

# --- as tres limpezas, nomeadas para o comparador poder cobra-las -----------
RUIDO_FONTE = ("Arial", 10.0)          # limpeza 1
LARGURA_CERTA = 4.0                    # limpeza 2 (o Sheets devolve 3.63)
CF_DE_FABRICA = "FFB7E1CD"             # limpeza 3

wb = load_workbook(ORIG)
estilos, indice = [], {}

def cor(c):
    if c is None or getattr(c, "rgb", None) in (None, "00000000"):
        return None
    rgb = c.rgb
    return rgb if isinstance(rgb, str) else None

def chave_estilo(cel):
    f, p, b, a = cel.font, cel.fill, cel.border, cel.alignment
    fonte = None
    if f and f.name:
        # limpeza 1: o Arial 10 de fabrica nao entra no estilo
        if not (f.name == RUIDO_FONTE[0] and f.sz == RUIDO_FONTE[1]):
            fonte = [f.name, f.sz, cor(f.color), bool(f.b), bool(f.i)]
    fundo = cor(p.start_color) if (p and p.fill_type == "solid") else None
    bordas = None
    if b:
        lados = {}
        for lado in ("top", "bottom", "left", "right"):
            s = getattr(b, lado)
            if s and s.style:
                lados[lado] = [s.style, cor(s.color)]
        bordas = lados or None
    alinha = None
    if a and (a.horizontal or a.vertical or a.wrap_text or a.text_rotation):
        alinha = [a.horizontal, a.vertical, bool(a.wrap_text), a.text_rotation or 0]
    fmt = cel.number_format if cel.number_format != "General" else None
    return json.dumps([fonte, fundo, bordas, alinha, fmt], ensure_ascii=False,
                      sort_keys=True)

def idx_estilo(cel):
    k = chave_estilo(cel)
    if k == '[null, null, null, null, null]':
        return None
    if k not in indice:
        indice[k] = len(estilos)
        estilos.append(json.loads(k))
    return indice[k]

abas = []
for nome in wb.sheetnames:
    s = wb[nome]
    celulas = []
    for lin in s.iter_rows():
        for c in lin:
            e = idx_estilo(c)
            if c.value is None and e is None:
                continue
            celulas.append([c.coordinate, c.value, e])

    # as colunas: o Sheets colapsa tudo num range so, e a largura volta a 4,0
    cols = []
    for k, v in s.column_dimensions.items():
        if v.width:
            larg = LARGURA_CERTA if abs(v.width - 3.63) < 0.01 else v.width
            cols.append([v.min, v.max, larg])
    linhas_h = [[int(k), v.height] for k, v in s.row_dimensions.items() if v.height]

    menus = []
    for v in s.data_validations.dataValidation:
        menus.append({"onde": str(v.sqref), "tipo": v.type, "formula": v.formula1,
                      "vazio_ok": bool(v.allow_blank),
                      "mostra_seta": not bool(v.showDropDown)})

    # a condicional: as de estado da A5 ficam, a verde de fabrica sai
    cfs = []
    for faixa in s.conditional_formatting:
        for r in faixa.rules:
            f_bg = None
            if r.dxf and r.dxf.fill and r.dxf.fill.bgColor:
                f_bg = cor(r.dxf.fill.bgColor)
            f_cor = None
            if r.dxf and r.dxf.font and r.dxf.font.color:
                f_cor = cor(r.dxf.font.color)
            if f_bg == CF_DE_FABRICA:          # limpeza 3
                continue
            # o Sheets exporta a formula entre aspas extras
            forms = [re.sub(r'^"|"$', "", x) for x in (r.formula or [])]
            cfs.append({"onde": str(faixa.sqref), "tipo": r.type,
                        "formula": forms, "cor_texto": f_cor, "fundo": f_bg})

    mescladas = [str(r) for r in s.merged_cells.ranges]

    imagens = []
    for i, im in enumerate(getattr(s, "_images", [])):
        arq = f"{nome.lower().replace(' ', '-')}-{i+1}.png"
        try:
            dados = im.ref.getvalue() if hasattr(im.ref, "getvalue") else open(im.ref, "rb").read()
            open(os.path.join(ARTE, arq), "wb").write(dados)
        except Exception as exc:
            print(f"  [aviso] nao extrai a imagem {i+1} de {nome}: {exc}")
            continue
        a = im.anchor
        imagens.append({"arquivo": arq, "col": a._from.col + 1, "lin": a._from.row + 1,
                        "larg": im.width, "alt": im.height})

    abas.append({"nome": nome, "estado": s.sheet_state,
                 "linhas": s.max_row, "colunas": s.max_column,
                 "colunas_larg": cols, "linhas_alt": linhas_h,
                 "altura_padrao": s.sheet_format.defaultRowHeight,
                 "grade": bool(s.sheet_view.showGridLines),
                 "celulas": celulas, "mescladas": mescladas,
                 "menus": menus, "condicional": cfs,
                 "imagens": imagens})

layout = {
    "_meta": {
        "o_que_e": "o desenho da Ficha (PROJETO M) 0.1, extraido do .xlsx que o "
                   "Mizuki mandou. O monta.py replica isto.",
        "origem": os.path.basename(ORIG),
        "veio_do_sheets": "https://docs.google.com/spreadsheets/d/"
                          "1rH43Xw6nneXwIPkI1VpsnPPkTZTPIqbiocY0KQdPwZ8/edit",
        "limpezas": [
            "as celulas em Arial 10 preto, todas vazias, saem: e o estilo de "
            "fabrica do Sheets. Quantas sao, o comparador conta -- numero em "
            "prosa deriva, e este ja derivou de 3165 para 3161 numa edicao.",
            "a largura 3,63 volta a 4,0, que e o que o gerador escreve antes da "
            "ida e volta de unidade",
            "a condicional verde B7E1CD de D114, o 'nao esta vazio' de fabrica do "
            "Google, sai",
        ],
        "onde_ela_vive": "Google Sheets. Por isso o IFS fica cru e o SPARKLINE "
                         "continua: no Excel os dois quebram, e isso esta aceito.",
    },
    "estilos": estilos,
    "abas": abas,
}
json.dump(layout, open(SAIDA, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

print(f"layout escrito: {SAIDA}")
print(f"  {len(estilos)} estilos distintos")
for a in abas:
    print(f"  {a['nome']:10} {len(a['celulas']):>5} celulas · "
          f"{len(a['mescladas']):>3} mescl. · {len(a['menus'])} menus · "
          f"{len(a['condicional'])} cond. · {len(a['imagens'])} imgs")
