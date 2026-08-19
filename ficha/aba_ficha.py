# -*- coding: utf-8 -*-
"""A aba FICHA: criacao e progressao. 46 colunas, ~1300 px.
Toda formula sai do manual; nenhum numero derivado e digitado pelo jogador."""
from estilo import *
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.datavalidation import DataValidation

COLS, GUT = 46, 1

def caixa(ws, c1, r1, c2, r2, rot, valor, pt=PT_VALOR, cor=TEXTO, fundo=PAINEL,
          nome=None, formato=None, alinha=None):
    """um campo: rotulo em cima, valor embaixo, tudo dentro de um bloco pintado"""
    pinta(ws, c1, r1, c2, r2, fundo)
    rotulo(ws, c1, r1, c2, rot)
    junta(ws, c1, r1 + 1, c2, r2)
    cel = escreve(ws, c1, r1 + 1, valor, pt=pt, cor=cor, nome=nome,
                  alinha=alinha or CENTRO, formato=formato)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            lados = ("t" if r == r1 else "") + ("b" if r == r2 else "") + \
                    ("l" if c == c1 else "") + ("r" if c == c2 else "")
            if lados:
                ws.cell(row=r, column=c).border = borda(LINHA, "thin", lados)
    return cel

def monta(wb, CAT, DEC, ref):
    ws = wb.create_sheet("FICHA")
    ws.sheet_view.showGridLines = False
    for c in range(1, COLS + 1):
        ws.column_dimensions[L(c)].width = LARG_COL
    for r in range(1, 120):
        ws.row_dimensions[r].height = ALT_LIN
    pinta(ws, 1, 1, COLS, 120, FUNDO)
    R = {}                                    # onde cada numero ficou

    def dv(formula, c1, r1, c2, r2):
        v = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(v)
        v.add(f"{L(c1)}{r1}:{L(c2)}{r2}")

    # ---------------------------------------------------------------- topo
    pinta(ws, 1, 1, COLS, 4, PAINEL_ALTO)
    junta(ws, 1, 1, 12, 2);  escreve(ws, 1, 1, "PROJETO M", nome=TITULO, pt=20, cor=OSSO, alinha=ESQ)
    junta(ws, 1, 3, 12, 3);  escreve(ws, 1, 3, "ficha de personagem", nome=TITULO,
                                     pt=PT_ROTULO, cor=TEXTO_FRACO, alinha=ESQ)
    junta(ws, 14, 1, 34, 2)
    R["nome"] = escreve(ws, 14, 1, None, pt=18, cor=TEXTO, nome=TITULO, alinha=ESQ)
    rotulo(ws, 14, 3, 34, "nome do personagem")
    # o aviso de versao, decisao A1
    junta(ws, 36, 1, COLS, 2)
    escreve(ws, 36, 1,
            f'=IF({ref["carimbo_local"]}={ref["carimbo_central"]},'
            f'"catálogo v"&{ref["carimbo_local"]}&" · em dia",'
            f'"⚠ ficha na v"&{ref["carimbo_local"]}&" · a atual é a v"&{ref["carimbo_central"]})',
            pt=PT_ROTULO, cor=TEXTO_FRACO)
    junta(ws, 36, 3, COLS, 3)
    escreve(ws, 36, 3, "呪術廻戦", nome=MARCA, pt=PT_ROTULO, cor=LINHA)

    # ------------------------------------------------------- identidade
    r = 6
    escreve(ws, 1, r, "QUEM É", nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    R["caminho"] = caixa(ws, 1, r, 10, r + 2, "caminho", None)
    dv(ref["Caminhos"], 1, r + 1, 10, r + 1)
    R["trilha"]  = caixa(ws, 12, r, 21, r + 2, "trilha", None)
    dv(ref["Trilhas"], 12, r + 1, 21, r + 1)
    R["origem"]  = caixa(ws, 23, r, 32, r + 2, "origem", None)
    dv(ref["Origens"], 23, r + 1, 32, r + 1)
    R["nivel"]   = caixa(ws, 34, r, 39, r + 2, "nível", 2, pt=16, nome=TITULO, cor=OSSO)
    R["xp"]      = caixa(ws, 41, r, COLS, r + 2, "xp (na mão)", 0, cor=TEXTO_FRACO)
    NIV = f"$" + L(34) + "$" + str(r + 1)

    # ------------------------------------------------------- atributos
    r += 4
    escreve(ws, 1, r, "ATRIBUTOS", nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    ATR = CAT["atributos"]["lista"]
    larg = 8
    for i, a in enumerate(ATR):
        c1 = 1 + i * (larg + 1)
        pinta(ws, c1, r, c1 + larg - 1, r + 5, PAINEL)
        rotulo(ws, c1, r, c1 + larg - 1, a[:3])
        junta(ws, c1, r + 1, c1 + larg - 1, r + 4)
        cel = escreve(ws, c1, r + 1, 0, pt=PT_GRANDE, nome=TITULO, cor=OSSO)
        junta(ws, c1, r + 5, c1 + larg - 1, r + 5)
        escreve(ws, c1, r + 5, a, pt=PT_ROTULO, cor=TEXTO_FRACO)
        R["atr_" + a] = f"${L(c1)}${r+1}"
        for rr in range(r, r + 6):
            for cc in range(c1, c1 + larg):
                lados = ("t" if rr == r else "") + ("b" if rr == r + 5 else "") + \
                        ("l" if cc == c1 else "") + ("r" if cc == c1 + larg - 1 else "")
                if lados:
                    ws.cell(row=rr, column=cc).border = borda(LINHA, "thin", lados)
    # pelo NOME, nunca pela posicao: a ordem do catalogo e Forca, Destreza,
    # Constituicao -- ler por indice fazia a Vida somar Destreza no lugar de Con
    FOR = R["atr_Força"]
    CON = R["atr_Constituição"]
    DES = R["atr_Destreza"]
    INT = R["atr_Inteligência"]
    ESS = R["atr_Essência"]

    # ------------------------------------------------------- reservas
    r += 7
    escreve(ws, 1, r, "AS TRÊS RESERVAS", nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    CAM = "$" + R["caminho"].coordinate.replace("$", "")
    CAM = "$" + CAM[1:2].replace("", "") + CAM[1:] if False else \
          f'${L(R["caminho"].column)}${R["caminho"].row}'
    TAB = ref["tabela_caminhos"]
    vida_max = (f'=IF({CAM}="","",(VLOOKUP({CAM},{TAB},3,FALSE)+{CON})'
                f'+(VLOOKUP({CAM},{TAB},4,FALSE)+{CON})*({NIV}-1))')
    pe_max   = f'=IF({CAM}="","",VLOOKUP({CAM},{TAB},5,FALSE)*{NIV})'
    int_max  = f'=20+8*({NIV}-1)'
    for i, (nome_r, f_max) in enumerate([("vida", vida_max), ("energia", pe_max),
                                          ("integridade", int_max)]):
        c1 = 1 + i * 16
        pinta(ws, c1, r, c1 + 14, r + 6, PAINEL)
        rotulo(ws, c1, r, c1 + 14, nome_r)
        # atual: numero grande, com a cor de estado (A5) por formatacao condicional
        junta(ws, c1, r + 1, c1 + 7, r + 5)
        atual = escreve(ws, c1, r + 1, 0, pt=PT_GRANDE, nome=TITULO, cor=OSSO)
        caixa(ws, c1 + 9, r + 1, c1 + 14, r + 2, "máximo", f_max, cor=TEXTO_FRACO, fundo=PAINEL_ALTO)
        caixa(ws, c1 + 9, r + 3, c1 + 14, r + 4, "temp", 0, cor=TEXTO_FRACO, fundo=PAINEL_ALTO)
        junta(ws, c1, r + 6, c1 + 7, r + 6)
        escreve(ws, c1, r + 6, "atual", pt=PT_ROTULO, cor=TEXTO_FRACO)
        # A4: a caixinha de delta, ao lado do atual que continua editavel
        caixa(ws, c1 + 9, r + 5, c1 + 14, r + 6, "±", None, cor=OSSO, fundo=PAINEL_ALTO)
        R[nome_r] = f"${L(c1)}${r+1}"
        R[nome_r + "_max"] = f"${L(c1+9)}${r+2}"
        R[nome_r + "_temp"] = f"${L(c1+9)}${r+4}"
        R[nome_r + "_delta"] = f"${L(c1+9)}${r+6}"

    # estagio de alma: a ficha le, o jogador nunca marca
    r += 7
    junta(ws, 33, r, COLS, r)
    ITG, ITGM = R["integridade"], R["integridade_max"]
    escreve(ws, 33, r,
            f'=IF({ITG}=""," ",IF({ITG}<=0,"estágio 4 · você não é mais você",'
            f'IF({ITG}<={ITGM}/4,"estágio 3 · desvantagem em ataques e TRs",'
            f'IF({ITG}<={ITGM}/2,"estágio 2 · metade do deslocamento, +1 PE por Classe",'
            f'IF({ITG}<={ITGM}*3/4,"estágio 1 · desvantagem em perícias","alma inteira")))))',
            pt=PT_ROTULO, cor=TEXTO_FRACO, alinha=ESQ)
    R["estagio_alma"] = f"${L(33)}${r}"

    # ------------------------------------------------------- derivados
    r += 2
    escreve(ws, 1, r, "OS NÚMEROS QUE CAEM SOZINHOS", nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    MAE = f'(1+COUNTIF({ref["m_maestria"]},"<="&{NIV}))'
    # C3: a protecao e escolha, nunca constante. Refino de graca = 1 + marcos.
    REF_GRACA = f'(1+COUNTIF({ref["m_marcos"]},"<="&{NIV}))'
    R["refino"] = None
    # onde cada campo cai, pela mesma conta que o laco usa embaixo
    pos = lambda i: (1 + (i % 5) * 9, r + (i // 5) * 4 + 1)
    C_PROT, R_PROT = pos(1)
    C_EQUI, R_EQUI = pos(2)
    EQUI = f'${L(C_EQUI)}${R_EQUI}'
    PROT = f'${L(C_PROT)}${R_PROT}'
    campos = [
        ("defesa",        f'=10+{DES}+{PROT}'),
        ("proteção",      f'=IF({EQUI}="",FLOOR({REF_GRACA}/3,1)+1,{EQUI})'),
        ("equipamento",   None),
        ("iniciativa",    f'="d20 + "&{DES}'),
        ("deslocamento",  "9 m"),
        ("maestria",      f'={MAE}'),
        ("cd de feitiço", f'=10+2+{MAE}'),
        ("conjuração",    f'="d20 + "&(2+{MAE})'),
        ("corpo a corpo", f'="d20 + "&{FOR}'),
        ("à distância",   f'="d20 + "&{DES}'),
    ]
    for i, (rot, val) in enumerate(campos):
        c1 = 1 + (i % 5) * 9
        rr = r + (i // 5) * 4
        alvo = caixa(ws, c1, rr, c1 + 7, rr + 2, rot, val,
                     pt=14, nome=TITULO, cor=OSSO if val else TEXTO)
        R[rot] = f"${L(c1)}${rr+1}"
    r += 8

    # ------------------------------------------------------- progressao
    escreve(ws, 1, r, "PROGRESSÃO · a ficha calcula, você só digita o nível",
            nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    prog = [
        ("espaços de feitiço", f'=2+INT({NIV}/2)+COUNTIF({ref["m_marcos"]},"<="&{NIV})'),
        ("refino de graça",    f'={REF_GRACA}'),
        ("classe máxima",      f'=COUNTIF({ref["m_classe"]},"<="&{NIV})'),
        ("classe 0 grátis",    f'=2+COUNTIF({ref["m_classe0"]},"<="&{NIV})'),
        ("perícias treinadas", f'=COUNTIF(${L(1)}${r+4}:${L(1)}${r+27},"x")'),
    ]
    for i, (rot, val) in enumerate(prog):
        c1 = 1 + i * 9
        caixa(ws, c1, r, c1 + 7, r + 2, rot, val, pt=14, nome=TITULO, cor=OSSO)
        R[rot] = f"${L(c1)}${r+1}"
    r += 4

    # ------------------------------------------------------- pericias
    escreve(ws, 1, r, "PERÍCIAS · marque com x as 8 (ou 9) treinadas",
            nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    linha0 = r
    for i, (nome_p, d) in enumerate(CAT["pericias"].items()):
        cc = 1 + (i // 12) * 16
        rr = r + (i % 12)
        pinta(ws, cc, rr, cc + 14, rr, PAINEL if (i % 2 == 0) else PAINEL_ALTO)
        escreve(ws, cc, rr, None, alinha=CENTRO, cor=OSSO)          # a marca
        junta(ws, cc + 1, rr, cc + 9, rr)
        escreve(ws, cc + 1, rr, nome_p, pt=PT_ROTULO, cor=TEXTO, alinha=ESQ)
        junta(ws, cc + 10, rr, cc + 12, rr)
        escreve(ws, cc + 10, rr, d["atributo"][:3], pt=PT_ROTULO, cor=TEXTO_FRACO)
        junta(ws, cc + 13, rr, cc + 14, rr)
        escreve(ws, cc + 13, rr,
                f'={R["atr_"+d["atributo"]]}+IF(${L(cc)}${rr}="",0,{MAE})',
                pt=PT_ROTULO, cor=OSSO)
    R["pericias_marca"] = (1, linha0, linha0 + 11)
    r += 12

    # ------------------------------------------------------- oficios e TRs
    r += 1
    escreve(ws, 1, r, "OFÍCIOS", nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    escreve(ws, 25, r, "TESTES DE RESISTÊNCIA", nome=TITULO, pt=PT_TITULO, cor=BLOCO, alinha=ESQ)
    r += 1
    for i, of in enumerate(CAT["oficios"]):
        cc, rr = 1 + (i // 6) * 12, r + (i % 6)
        pinta(ws, cc, rr, cc + 10, rr, PAINEL if i % 2 == 0 else PAINEL_ALTO)
        escreve(ws, cc, rr, None, cor=OSSO)
        junta(ws, cc + 1, rr, cc + 10, rr)
        escreve(ws, cc + 1, rr, of, pt=PT_ROTULO, alinha=ESQ)
    TRS = [t for t, v in CAT["testes_de_resistencia"].items() if isinstance(v, dict)]
    BON = CAT["testes_de_resistencia"]["bonus_se_treinado"]
    for i, t in enumerate(TRS):
        rr = r + i
        pinta(ws, 25, rr, COLS, rr, PAINEL if i % 2 == 0 else PAINEL_ALTO)
        escreve(ws, 25, rr, None, cor=OSSO)
        junta(ws, 26, rr, 34, rr)
        escreve(ws, 26, rr, t, pt=PT_ROTULO, alinha=ESQ)
        atr = CAT["testes_de_resistencia"][t]["atributo"]
        junta(ws, 35, rr, 41, rr)
        escreve(ws, 35, rr, " ou ".join(a[:3] for a in atr), pt=PT_ROTULO, cor=TEXTO_FRACO)
        junta(ws, 42, rr, COLS, rr)
        escreve(ws, 42, rr,
                f'={R["atr_"+atr[0]]}+IF(${L(25)}${rr}="",0,{BON})',
                pt=PT_ROTULO, cor=OSSO)
    return R
