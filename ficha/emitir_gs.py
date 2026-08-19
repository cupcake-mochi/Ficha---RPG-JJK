# -*- coding: utf-8 -*-
"""Emite o Apps Script que constroi a ficha DENTRO do Google Sheets.

Por que isto existe: o caminho pelo .xlsx morreu numa prova. O registro do
script devolveu 'imagens: 0' -- imagem que vem da importacao do .xlsx nao e
visivel pela API do Sheets, entao nao da nem para consertar o tamanho dela.
E ela nao era a unica coisa que a conversao quebrava: fonte, altura de linha,
caixa de selecao e o fundo das celulas mescladas quebravam tambem.

Aqui a planilha nasce nativa. Nada e traduzido, entao nada se perde na traducao.

O emissor NAO redesenha nada: ele le a mesma pasta de trabalho que o monta.py
ja produz, celula a celula, e vira instrucao. Layout e formula continuam
sendo os mesmos que os dez validadores conferem.
"""
import base64, json, os
from openpyxl.utils import get_column_letter as L

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
ARTE = os.path.join(RAIZ, "arte")
SAIDA = os.path.join(RAIZ, "apps-script", "Ficha.gs")

# a textura sai: 600 KB de base64 para um ruido que nao aparece na tela
ARTE_FORA = {"textura.png"}

def _cor(c):
    if c is None or getattr(c, "rgb", None) is None or not isinstance(c.rgb, str):
        return None
    h = c.rgb[-6:].upper()
    return None if h in ("000000",) else "#" + h

def _linha_alta(pts):
    """altura em PIXEL, medida pela maior letra da linha.

    O .xlsx guarda altura em ponto e o Sheets le em pixel; foi por isso que o
    'd20 + 0' aparecia cortado pela metade. 1 pt = 1.333 px, mais folga.
    """
    return max(21, int(max(pts) * 1.34) + 7) if pts else 21

def emitir(wb, ordem):
    abas = []
    for nome in ordem:
        ws = wb[nome]
        vals, estilos, chaves, fundos, bordas, merges = [], [], {}, [], [], []
        alturas, max_c, max_r = {}, 0, 0
        for linha in ws.iter_rows():
            for c in linha:
                if c.__class__.__name__ == "MergedCell":
                    continue
                r, col = c.row, c.column
                f, a = c.font, c.alignment
                pintado = _cor(c.fill.start_color) if c.fill and c.fill.fill_type else None
                if c.value is None and pintado is None and (not f or not f.name):
                    continue
                max_c, max_r = max(max_c, col), max(max_r, r)
                if c.value is not None:
                    # formula vai separada: setValues trata o texto como digitado
                    # pelo usuario, e ai a pontuacao segue o idioma da planilha.
                    # Numa planilha em portugues, COUNTIF(a,b) vira erro de
                    # analise. O setFormula sempre usa a notacao americana.
                    vals.append([r, col, c.value])
                    alturas.setdefault(r, []).append(f.size or 11)
                if pintado:
                    fundos.append([r, col, pintado])   # comprimido depois, em faixas
                if f and f.name:
                    ch = (f.name, f.size, _cor(f.color), bool(f.bold),
                          a.horizontal or "left", a.vertical or "middle",
                          int(a.textRotation or 0))
                    if ch not in chaves:
                        chaves[ch] = len(estilos)
                        estilos.append(list(ch))
                    estilos_id = chaves[ch]
                    vals[-1].append(estilos_id) if (c.value is not None) else None
                    if c.value is None:
                        continue
                if c.border and c.border.top and c.border.top.style:
                    bordas.append([r, col, _cor(c.border.top.color) or "#493F54"])
        # o fundo em faixas: 15 mil celulas pintadas viravam 15 mil entradas.
        # Vizinhas da mesma cor na mesma linha viram uma faixa so, e a cor de
        # base nem entra -- o script ja comeca com ela.
        BASE = "#120F1D"
        por_linha = {}
        for r, c, cor in fundos:
            if cor != BASE:
                por_linha.setdefault(r, {})[c] = cor
        faixas = []
        for r in sorted(por_linha):
            cols = sorted(por_linha[r])
            ini = ant = cols[0]
            for c in cols[1:] + [None]:
                if c is None or c != ant + 1 or por_linha[r][c] != por_linha[r][ini]:
                    faixas.append([r, ini, ant, por_linha[r][ini]])
                    if c is not None:
                        ini = c
                ant = c if c is not None else ant
        fundos = faixas

        for m in ws.merged_cells.ranges:
            merges.append([m.min_row, m.min_col, m.max_row, m.max_col])
            max_c, max_r = max(max_c, m.max_col), max(max_r, m.max_row)
        dv = []
        for v in ws.data_validations.dataValidation:
            if v.type == "list" and v.formula1:
                for rg in str(v.sqref).split():
                    dv.append([rg, v.formula1.replace("DADOS!", "DADOS!")])
        import estilo
        imgs = [[i["lin"], i["col"], i["larg"], i["alt"], i["nome"]]
                for i in estilo.COLOCADAS
                if i["aba"] == nome and i["nome"] not in ARTE_FORA]
        abas.append({
            "nome": nome, "cols": max(max_c, 12), "rows": max_r + 2,
            "larg": int(round((ws.column_dimensions["A"].width or 4.0) * 7)),
            "vals": vals, "estilos": estilos, "fundos": fundos, "bordas": bordas,
            "merges": merges, "dv": dv, "imgs": imgs,
            "alturas": {str(r): _linha_alta(p) for r, p in alturas.items()},
            "oculta": ws.sheet_state == "hidden",
        })
    arte = {}
    for f in sorted(os.listdir(ARTE)):
        if f.endswith(".png") and f not in ARTE_FORA and "contato" not in f:
            arte[f] = base64.b64encode(open(os.path.join(ARTE, f), "rb").read()).decode()
    return abas, arte

def escrever(wb, ordem, caixas=None):
    abas, arte = emitir(wb, ordem)
    for a in abas:
        a["caixas"] = caixas if (caixas and a["nome"] == "FICHA") else []
    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    corpo = open(os.path.join(AQUI, "modelo.gs.js"), encoding="utf-8").read()
    with open(SAIDA, "w", encoding="utf-8") as fp:
        fp.write("// GERADO POR ficha/emitir_gs.py — não edite este arquivo na mão.\n")
        fp.write("// O layout e as fórmulas moram no gerador Python, que os dez\n")
        fp.write("// validadores conferem. Aqui é só o transporte.\n\n")
        fp.write("var ABAS = " + json.dumps(abas, ensure_ascii=False, separators=(",", ":")) + ";\n\n")
        fp.write("var ARTE = " + json.dumps(arte, ensure_ascii=False,
                                    separators=(",", ":")) + ";\n\n")
        fp.write(corpo)
    return SAIDA, sum(len(a["vals"]) for a in abas), len(arte)
