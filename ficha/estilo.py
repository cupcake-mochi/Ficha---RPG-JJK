# -*- coding: utf-8 -*-
"""O vocabulario visual da ficha do Projeto M.

Nada de cor, fonte ou tamanho esta escolhido aqui: tudo le do
decisoes-ficha.json, que e o dono das decisoes A5 (cor de estado), C4 (fontes),
C5 (arte) e C6 (a ficha e um documento).

A gramatica, em uma frase: REGUA no lugar de CAIXA. Um fio embaixo do valor,
nunca borda em volta dele -- e isso, mais que qualquer outra coisa, que separa
desenho de formulario.
"""
import json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as _Img
from openpyxl.utils import get_column_letter as L

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARTE = os.path.join(RAIZ, "arte")
DEC  = json.load(open(os.path.join(RAIZ, "decisoes-ficha.json"), encoding="utf-8"))

# --- paleta --------------------------------------------------------------
FUNDO, PAINEL, PAINEL_ALTO = "120F1D", "211C35", "30294D"
LINHA, BLOCO, TEXTO_FRACO  = "493F54", "756588", "998BA9"
TEXTO, TINTA, PAPEL        = "F4F1F7", "0A0810", "17131F"

_G = {g["nome"]: g["hex"] for g in DEC["A5_acento"]["degraus"]}
OSSO, AMBAR, VERMELHO = _G["osso"], _G["ambar"], _G["vermelho"]

# --- fontes, decisao C4 --------------------------------------------------
_F = DEC["C4_fontes"]
CORPO      = _F["corpo"]["fonte"]        # Roboto     · dado, e sozinha na MESA
TITULO     = _F["titulo"]["fonte"]       # Oswald     · rotulo e numero de secao
DOCUMENTO  = _F["documento"]["fonte"]    # Shippori   · nome e valor de campo
SERIE      = _F["serie"]["fonte"]        # Courier    · registro e data
MARCA      = _F["marca"]["fonte"]        # Yuji Syuku · kanji
KANJI_PISO = _F["kanji_piso_pt"]         # 22: abaixo disso o traco funde

# a Castoro tem altura de x 0.475, a menor das candidatas: o corpo sobe um
# ponto para compensar. O numero saiu da medida, nao do olho.
PT_ROTULO, PT_VALOR, PT_TITULO, PT_GRANDE = 9, 12, 12, 26
LARG_COL, ALT_LIN = 4.0, 15

def fill(c): return PatternFill("solid", start_color=c, end_color=c)

def pinta(ws, c1, r1, c2, r2, cor):
    f = fill(cor)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = f

def junta(ws, c1, r1, c2, r2):
    if (c1, r1) != (c2, r2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def txt(ws, c, r, v, nome=None, pt=PT_VALOR, cor=TEXTO, neg=False,
        al="left", va="center", rot=0, fundo=None, ate=None, fmt=None):
    if ate:
        junta(ws, c, r, ate[0], ate[1])
    cel = ws.cell(row=r, column=c, value=v)
    if nome == MARCA and pt < KANJI_PISO:
        raise ValueError(f"kanji a {pt} pt: o piso medido e {KANJI_PISO} pt (decisao C4)")
    cel.font = Font(name=nome or CORPO, size=pt, color=cor, bold=neg)
    cel.alignment = Alignment(horizontal=al, vertical=va, textRotation=rot)
    if fundo: cel.fill = fill(fundo)
    if fmt:   cel.number_format = fmt
    return cel

def regua(ws, c1, r, c2, cor=LINHA, grossa=False):
    """a linha horizontal que substitui a caixa. E o movimento principal."""
    s = Side(style="medium" if grossa else "thin", color=cor)
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).border = Border(top=s)

def campo(ws, c1, r, larg, rot, valor, pt=16, cor=TEXTO, nome=None, fonte_rot=None):
    """rotulo pequeno em cima, valor grande embaixo, regua debaixo. Sem caixa."""
    txt(ws, c1, r, rot.upper(), nome=fonte_rot or TITULO, pt=8, cor=TEXTO_FRACO,
        ate=(c1 + larg, r))
    cel = txt(ws, c1, r + 1, valor, nome=nome or DOCUMENTO, pt=pt, cor=cor,
              ate=(c1 + larg, r + 1))
    regua(ws, c1, r + 2, c1 + larg, LINHA)
    return cel

def secao(ws, r, num, nome, c1=4, c2=24):
    """o numero da secao entra grande e apagado: ornamento que tambem orienta"""
    txt(ws, c1, r, num, nome=TITULO, pt=22, cor=PAINEL_ALTO, ate=(c1 + 2, r + 1))
    txt(ws, c1 + 3, r, nome, nome=TITULO, pt=PT_TITULO, cor=BLOCO, ate=(c2, r))
    return r + 2

# onde cada imagem foi parar. O emissor le daqui em vez de vasculhar o
# openpyxl: o ancoradouro dele muda de forma conforme como voce o define.
COLOCADAS = []

def arte(ws, nome, col, lin, larg_px, alt_px=None):
    """imagem flutuante. So em area morta: ela bloqueia a digitacao embaixo."""
    caminho = os.path.join(ARTE, nome)
    if not os.path.exists(caminho):
        return None
    im = _Img(caminho)
    proporcao = im.height / im.width
    im.width = larg_px
    im.height = alt_px if alt_px else int(larg_px * proporcao)
    im.anchor = ws.cell(row=lin, column=col).coordinate
    ws.add_image(im)
    COLOCADAS.append({"aba": ws.title, "nome": nome, "col": col, "lin": lin,
                      "larg": im.width, "alt": im.height})
    return im

def lombada(ws, linhas, texto_topo, texto_baixo):
    """a faixa escura na lateral, com texto girado. Planilha nenhuma tem isso."""
    pinta(ws, 1, 1, 2, linhas, TINTA)
    txt(ws, 1, 8, texto_topo, nome=TITULO, pt=11, cor=LINHA, rot=90,
        al="center", ate=(2, 26))
    txt(ws, 1, 30, texto_baixo, nome=TITULO, pt=11, cor=LINHA, rot=90,
        al="center", ate=(2, 48))

def barra(cel_atual, cel_max, cor_formula):
    """SPARKLINE nativo. Funciona no app de celular -- o Mizuki confirmou.

    O IFERROR nao e enfeite: com o Caminho vazio o maximo vira texto vazio, a
    divisao da cor estoura, e a ficha em branco abre com #DIV/0! na cara do
    jogador. Ficha vazia tem que abrir limpa.
    """
    return (f'=IFERROR(SPARKLINE({cel_atual},{{"charttype","bar";"max",{cel_max};'
            f'"color1",{cor_formula}}}),"")')

def cor_de_estado(atual, maximo):
    """a formula da decisao A5, em texto, para entrar dentro do SPARKLINE"""
    # N() resolve o campo vazio: sem ele "" nao e 0 e a divisao estoura
    return (f'IF(N({maximo})=0,"#{OSSO}",'
            f'IF({atual}/{maximo}<=0.25,"#{VERMELHO}",'
            f'IF({atual}/{maximo}<=0.5,"#{AMBAR}","#{OSSO}")))')

def base(wb, nome, cols=46, linhas=80):
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False
    for c in range(1, cols + 1):
        ws.column_dimensions[L(c)].width = LARG_COL
    for r in range(1, linhas + 1):
        ws.row_dimensions[r].height = ALT_LIN
    pinta(ws, 1, 1, cols, linhas, FUNDO)
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = ws.page_setup.fitToHeight = 1
    return ws


# --- compatibilidade com as abas que ainda falam a API antiga -------------
# A DADOS e a QUEM E nao mudaram de desenho, entao nao vale reescrever as duas
# so para trocar o nome das funcoes. Estes apelidos somem quando elas mudarem.
CENTRO   = Alignment("center", "center")
CENTRO_Q = Alignment("center", "center", wrap_text=True)
ESQ      = Alignment("left", "center", indent=1)
ESQ_Q    = Alignment("left", "top", wrap_text=True, indent=1)

def fonte(nome=None, pt=PT_VALOR, cor=TEXTO, negrito=False, italico=False):
    return Font(name=nome or CORPO, size=pt, color=cor, bold=negrito, italic=italico)

def borda(cor=LINHA, estilo="thin", lados="tblr"):
    s, n = Side(style=estilo, color=cor), Side(style=None)
    return Border(top=s if "t" in lados else n, bottom=s if "b" in lados else n,
                  left=s if "l" in lados else n, right=s if "r" in lados else n)

def escreve(ws, c, r, valor, **kw):
    cel = ws.cell(row=r, column=c, value=valor)
    cel.font = fonte(kw.get("nome"), kw.get("pt", PT_VALOR), kw.get("cor", TEXTO),
                     kw.get("negrito", False), kw.get("italico", False))
    cel.alignment = kw.get("alinha", CENTRO)
    if kw.get("fundo"):   cel.fill = fill(kw["fundo"])
    if kw.get("borda"):   cel.border = kw["borda"]
    if kw.get("formato"): cel.number_format = kw["formato"]
    return cel

def rotulo(ws, c1, r, c2, texto, cor=TEXTO_FRACO, nome=None):
    junta(ws, c1, r, c2, r)
    return escreve(ws, c1, r, texto.upper(), nome=nome or TITULO, pt=PT_ROTULO, cor=cor)
