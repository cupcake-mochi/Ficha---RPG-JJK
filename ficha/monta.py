# -*- coding: utf-8 -*-
"""Ficha do Projeto M em planilha, feita para NAO parecer planilha.
Tecnica: coluna de ~18px, gridlines desligadas, tudo desenhado por merge.
A celula vira pixel. Tres abas visiveis, uma oculta com os catalogos."""
import json, sys
sys.path.insert(0, '.')
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from estilo import *

CAT = json.load(open('../catalogo-projeto-m.json', encoding='utf-8'))
KAORI = {"Força":3, "Constituição":2, "Destreza":2, "Inteligência":1, "Essência":1}
TREINADAS = ["Atletismo","Intuição","História","Hierarquia","Sobrevivência",
             "Sentir Energia","Percepção","Intimidação"]
wb = Workbook()

def nova(titulo, ncol=52, nlin=86):
    ws = wb.create_sheet(titulo) if titulo != "FICHA" else wb.active
    ws.title = titulo
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMEIXA
    for c in range(1, ncol+1):
        letra = chr(64+c) if c <= 26 else 'A'+chr(64+c-26)
        ws.column_dimensions[letra].width = 2.35
    for r in range(1, nlin+1):
        ws.row_dimensions[r].height = 14.5
    pinta(ws, "A1", f"AZ{nlin}", FUNDO)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.15
    ws.page_margins.top = ws.page_margins.bottom = 0.15
    ws.print_area = f"A1:AZ{nlin}"
    return ws

def faixa(ws, ref, txt, cor=AMEIXA_ESC, tcor=OSSO, sz=7):
    """cabecalho de secao: barra fina com o titulo espacado"""
    caixa(ws, ref, " ".join(txt), cor, Font(name=TITULO, size=sz, bold=True, color=tcor), ESQ)

def campo(ws, ref_rot, ref_val, rot, valor_=None, cor_val=PAINEL2, sz=11, cor_txt=OSSO, borda_=AMEIXA_ESC):
    caixa(ws, ref_rot, rot, None, rotulo())
    return caixa(ws, ref_val, valor_, cor_val, valor(sz, cor_txt), CENTRO, borda_=borda_)

# ============================================================ ABA 1 · FICHA
ws = nova("FICHA")
caixa(ws, "B2:AY5", None, AMEIXA_ESC, merge=False)
caixa(ws, "B2:D5", "呪", AMEIXA, Font(name=TITULO, size=22, bold=True, color=OSSO))
caixa(ws, "E2:T3", "  PROJETO M", None, Font(name=TITULO, size=13, bold=True, color=OSSO), ESQ)
caixa(ws, "E4:T5", "  RPG DE MESA DE JUJUTSU KAISEN", None, Font(name=CORPO, size=6, color=AMEIXA_CLR), ESQ)
caixa(ws, "V2:AL5", "F I C H A   D E   P E R S O N A G E M", None,
      Font(name=CORPO, size=6, color=TEXTO_FRACO), CENTRO)
caixa(ws, "AN2:AY5", "NÍVEL 2  ·  GRAU 4", AMEIXA, Font(name=TITULO, size=9, bold=True, color=OSSO), CENTRO)

caixa(ws, "B7:AA8", "PERSONAGEM", None, rotulo(AMEIXA_CLR))
caixa(ws, "B9:AA12", "Kaori", AMEIXA, Font(name=TITULO, size=18, bold=True, color=OSSO), ESQ, borda_=AMEIXA_CLR)
campo(ws, "AC7:AK8", "AC9:AK12", "JOGADOR", "")
campo(ws, "AM7:AR8", "AM9:AR12", "PATENTE", "Grau 4", sz=10)
campo(ws, "AT7:AY8", "AT9:AY12", "NÍVEL", 2, cor_val=AMEIXA, sz=15)

# --- atributos
faixa(ws, "B14:M15", "ATRIBUTOS")
for i, a in enumerate(CAT["atributos"]["lista"]):
    l = 16 + i*4
    caixa(ws, f"B{l}:H{l+2}", f" {a.upper()}", PAINEL, Font(name=TITULO, size=7, bold=True, color=TEXTO_FRACO), ESQ)
    caixa(ws, f"I{l}:M{l+2}", KAORI[a], PAINEL2, valor(15, OSSO), CENTRO, borda_=AMEIXA)
caixa(ws, "B36:M37", "9 PONTOS · NENHUM ACIMA DE 3", None, Font(name=CORPO, size=5, color=TEXTO_FRACO))

# --- testes de resistencia
faixa(ws, "B39:M40", "RESISTÊNCIA")
for i, (t, d) in enumerate([(k,v) for k,v in CAT["testes_de_resistencia"].items() if isinstance(v, dict)]):
    l = 41 + i*3
    tr = "■" if t in ("Físico","Vigor") else "☐"
    caixa(ws, f"B{l}:C{l+1}", tr, PAINEL, Font(name=CORPO, size=8, color=AMEIXA_CLR), CENTRO)
    caixa(ws, f"D{l}:H{l+1}", t, PAINEL, corpo(7, OSSO), ESQ)
    caixa(ws, f"I{l}:M{l+1}", d["atributo"][0][:3].upper(), PAINEL2, corpo(6, TEXTO_FRACO), CENTRO)
caixa(ws, "B54:M55", "■ TREINADO · SOMA +2", None, Font(name=CORPO, size=5, color=TEXTO_FRACO))

# --- os numeros
faixa(ws, "O14:AA15", "OS NÚMEROS")
caixa(ws, "O16:T17", "VIDA", None, rotulo(SANGUE, 6))
caixa(ws, "O18:T23", "=(INDEX(caminho_vida1,MATCH($AM$16,caminho_nome,0))+$I$24)+(INDEX(caminho_vidaN,MATCH($AM$16,caminho_nome,0))+$I$24)*($AT$9-1)",
      PAINEL2, Font(name=TITULO, size=22, bold=True, color=SANGUE), CENTRO, borda_=SANGUE)
caixa(ws, "V16:AA17", "ENERGIA · PE", None, rotulo(ENERGIA, 6))
caixa(ws, "V18:AA23", "=INDEX(caminho_peN,MATCH($AM$16,caminho_nome,0))*$AT$9",
      PAINEL2, Font(name=TITULO, size=22, bold=True, color=ENERGIA), CENTRO, borda_=ENERGIA)
menores = [("DEFESA","=10+$I$20+1"), ("INTEGRIDADE","=20+8*($AT$9-1)"),
           ("INICIATIVA",'="d20 +"&$I$20'), ("DESLOCAMENTO",'="9 m"'),
           ("MAESTRIA","=1+(($AT$9>=10)+($AT$9>=18)+($AT$9>=26))"), ("CD DE FEITIÇO","=12+1+(($AT$9>=10)+($AT$9>=18)+($AT$9>=26))")]
for i, (rot, f) in enumerate(menores):
    col, c2 = ("O","T") if i % 2 == 0 else ("V","AA")
    l = 25 + (i//2)*6
    campo(ws, f"{col}{l}:{c2}{l+1}", f"{col}{l+2}:{c2}{l+4}", rot, f, sz=12)
faixa(ws, "O44:AA45", "ATAQUES")
for i, (rot, f) in enumerate([("CONJURAÇÃO",'="d20 +"&(2+1+(($AT$9>=10)+($AT$9>=18)+($AT$9>=26)))'),
                              ("CORPO A CORPO",'="d20 +"&$I$16'), ("À DISTÂNCIA",'="d20 +"&$I$20')]):
    l = 46 + i*3
    caixa(ws, f"O{l}:V{l+2}", " "+rot, PAINEL, Font(name=TITULO, size=7, bold=True, color=TEXTO_FRACO), ESQ)
    caixa(ws, f"W{l}:AA{l+2}", f, PAINEL2, valor(11, OSSO), CENTRO, borda_=AMEIXA_ESC)

# --- o que ele e
faixa(ws, "AC14:AY15", "O QUE ELE É")
for i, (rot, ref) in enumerate([("CAMINHO","AM16"), ("TRILHA","AM20"), ("ORIGEM","AM24")]):
    l = 16 + i*4
    caixa(ws, f"AC{l}:AK{l+2}", " "+rot, PAINEL, Font(name=TITULO, size=8, bold=True, color=TEXTO_FRACO), ESQ)
    caixa(ws, f"AM{l}:AY{l+2}", "", PAINEL2, corpo(9, OSSO), ESQ, borda_=AMEIXA)
ws["AM16"] = "Bastião"; ws["AM20"] = "Muro"; ws["AM24"] = "Latente"

# --- pericias
faixa(ws, "AC29:AY30", "PERÍCIAS   ·   23, E VOCÊ TREINA 8 OU 9")
por_attr = {}
for nome, dd in CAT["pericias"].items(): por_attr.setdefault(dd["atributo"], []).append(nome)
l = 31
for attr in CAT["atributos"]["lista"]:
    ps = por_attr.get(attr, [])
    if not ps: continue
    caixa(ws, f"AC{l}:AY{l}", " " + attr.upper(), None, Font(name=TITULO, size=6, bold=True, color=AMEIXA_CLR), ESQ)
    l += 1
    for i, p in enumerate(ps):
        col, c2 = [("AC","AK"),("AL","AT"),("AU","AY")][i % 3]
        ll = l + (i//3)
        marca = "■" if p in TREINADAS else "☐"
        caixa(ws, f"{col}{ll}:{c2}{ll}", f" {marca}  {p}", PAINEL if (i//3) % 2 == 0 else PAINEL2,
              corpo(6, OSSO if p in TREINADAS else TEXTO_FRACO), ESQ)
    l += (len(ps)+2)//3 + 1
faixa(ws, f"AC{l}:AY{l+1}", "OFÍCIOS   ·   11, E VOCÊ TREINA 2 OU 3")
l += 2
for i, o in enumerate(CAT["oficios"]):
    col, c2 = [("AC","AK"),("AL","AT"),("AU","AY")][i % 3]
    ll = l + (i//3)
    caixa(ws, f"{col}{ll}:{c2}{ll}", f" ☐  {o}", PAINEL if (i//3) % 2 == 0 else PAINEL2, corpo(6, TEXTO_FRACO), ESQ)
print(f"aba FICHA montada ate a linha {l+4}")

# ============================================================ ABA 2 · TÉCNICA
t = nova("TÉCNICA")
caixa(t, "B2:AY5", None, AMEIXA_ESC, merge=False)
caixa(t, "B2:D5", "呪", AMEIXA, Font(name=TITULO, size=22, bold=True, color=OSSO))
caixa(t, "E2:T3", "  O FUNDAMENTO", None, Font(name=TITULO, size=13, bold=True, color=OSSO), ESQ)
caixa(t, "E4:T5", "  A SUA TÉCNICA INATA", None, Font(name=CORPO, size=6, color=AMEIXA_CLR), ESQ)
caixa(t, "AN2:AY5", "=FICHA!B9", None, Font(name=TITULO, size=11, bold=True, color=AMEIXA_CLR),
      Alignment("right","center",indent=1))

caixa(t, "B7:AY8", "A REGRA   ·   UMA FRASE, VERIFICÁVEL PELA MESA, SEM NÚMERO", None, rotulo(AMEIXA_CLR))
caixa(t, "B9:AY12", "Tudo que eu prendo entre as minhas mãos fica mais pesado.", AMEIXA,
      Font(name=CORPO, size=10, italic=True, color=OSSO), Alignment("left","center",wrap_text=True,indent=2), borda_=AMEIXA_CLR)
caixa(t, "B14:AA15", "SELO   ·   O GESTO OU A CONDIÇÃO OBRIGATÓRIA", None, rotulo(AMEIXA_CLR, 5))
caixa(t, "B16:AA19", "As duas mãos precisam se tocar antes.", PAINEL, corpo(8, OSSO),
      Alignment("left","center",wrap_text=True,indent=1), borda_=AMEIXA_ESC)
caixa(t, "AC14:AY15", "PASSIVA LIVRE   ·   UMA, DE GRAÇA", None, rotulo(AMEIXA_CLR, 5))
caixa(t, "AC16:AY19", "", PAINEL, corpo(8, OSSO), Alignment("left","center",wrap_text=True,indent=1), borda_=AMEIXA_ESC)

faixa(t, "B21:AY22", "FAMÍLIAS   ·   DUAS LIVRES, TRÊS FECHADAS, QUATRO NO PREÇO NORMAL")
LIVRES, FECHADAS = ["Controle","Castigo"], ["Área","Auxiliares","Amparo"]
caixa(t, "B23:M23", "FAMÍLIA", None, rotulo(sz=5), ESQ)
caixa(t, "O23:S23", "LIVRE", None, rotulo(sz=5))
caixa(t, "U23:Y23", "FECHADA", None, rotulo(sz=5))
for i, (f, desc) in enumerate(CAT["familias"].items()):
    l = 24 + i*2
    livre, fech = f in LIVRES, f in FECHADAS
    cor = AMEIXA_ESC if livre else (PAINEL2 if fech else PAINEL)
    caixa(t, f"B{l}:M{l+1}", f" {f}", cor, corpo(8, OSSO if livre else TEXTO), ESQ)
    caixa(t, f"O{l}:S{l+1}", "■" if livre else "☐", cor, Font(name=CORPO, size=9, color=ENERGIA if livre else TEXTO_FRACO), CENTRO)
    caixa(t, f"U{l}:Y{l+1}", "■" if fech else "☐", cor, Font(name=CORPO, size=9, color=SANGUE if fech else TEXTO_FRACO), CENTRO)
    caixa(t, f"AA{l}:AY{l+1}", " " + desc, cor, corpo(6, TEXTO_FRACO), ESQ)
caixa(t, "B42:AY43", "LIVRE: A MELHORIA CUSTA METADE DA CLASSE A MENOS, MÍNIMO 1   ·   FECHADA: VOCÊ NUNCA COMPRA NADA DELA",
      None, Font(name=CORPO, size=5, color=TEXTO_FRACO), ESQ)

faixa(t, "B45:AY46", "FEITIÇOS   ·   NÍVEL 2: CLASSE 1, TRÊS CONHECIDOS, MAIS DOIS DE CLASSE 0")
cab = [("B","N","NOME DO FEITIÇO"), ("P","V","FORMA"), ("X","AB","PONTOS"), ("AD","AH","PE"), ("AJ","AY","MELHORIAS E RESTRIÇÕES")]
for c1, c2, h in cab: caixa(t, f"{c1}47:{c2}47", h, None, rotulo(sz=5), ESQ if h[0] in "NM" else CENTRO)
for n in range(3):
    l = 48 + n*3
    for c1, c2, h in cab:
        alv = ESQ if h.startswith(("NOME","MELHOR")) else CENTRO
        v = 3 if h in ("PONTOS","PE") else ""
        caixa(t, f"{c1}{l}:{c2}{l+1}", v, PAINEL if n % 2 == 0 else PAINEL2, corpo(7, OSSO), alv, borda_=AMEIXA_ESC)
caixa(t, "B58:AY59", "CLASSE 0   ·   DOIS, GRÁTIS, NÃO OCUPAM ESPAÇO", None, rotulo(AMEIXA_CLR, 5))
for n in range(2):
    caixa(t, f"B{60+n*2}:AY{61+n*2}", "", PAINEL if n % 2 == 0 else PAINEL2, corpo(7, OSSO), ESQ, borda_=AMEIXA_ESC)
caixa(t, "B65:AY67",
      "  ORÇAMENTO = 3 × CLASSE   ·   PE = 3 × CLASSE   ·   TETO DE DANO = 4 × CLASSE SOMANDO ALVOS E REPETIÇÕES\n"
      "  RESTRIÇÃO PAGA MELHORIA E NUNCA VIRA DADO   ·   DEVOLVE NO MÁXIMO 2 × CLASSE   ·   PONTO NÃO GASTO VIRA 1d8",
      AMEIXA_ESC, Font(name=CORPO, size=6, color=OSSO), Alignment("left","center",wrap_text=True,indent=1))
print("aba TÉCNICA montada")

# ============================================================ ABA 3 · QUEM É
q = nova("QUEM É")
caixa(q, "B2:AY5", None, AMEIXA_ESC, merge=False)
caixa(q, "B2:D5", "呪", AMEIXA, Font(name=TITULO, size=22, bold=True, color=OSSO))
caixa(q, "E2:T3", "  QUEM É ESSA PESSOA", None, Font(name=TITULO, size=13, bold=True, color=OSSO), ESQ)
caixa(q, "E4:T5", "  NADA AQUI ROLA DADO", None, Font(name=CORPO, size=6, color=AMEIXA_CLR), ESQ)
caixa(q, "AN2:AY5", "=FICHA!B9", None, Font(name=TITULO, size=11, bold=True, color=AMEIXA_CLR),
      Alignment("right","center",indent=1))
blocos = [("APARÊNCIA", "B", "AA", 7, 8), ("HISTÓRIA", "AC", "AY", 7, 8),
          ("O QUE A ORIGEM TE DEU", "B", "AA", 17, 5), ("O TRAÇO", "AC", "AY", 17, 5)]
for rot, col1, col2, l, alt in blocos:
    caixa(q, f"{col1}{l}:{col2}{l+1}", rot, None, rotulo(AMEIXA_CLR, 5))
    caixa(q, f"{col1}{l+2}:{col2}{l+alt}", "", PAINEL, corpo(8, OSSO),
          Alignment("left","top",wrap_text=True,indent=1), borda_=AMEIXA_ESC)
faixa(q, "B24:AY25", "OS DOIS LEGADOS   ·   UM DESTRANCA OBRIGATÓRIO, MAIS UM DE QUALQUER FORMATO")
for i, (rot, fmt) in enumerate([("LEGADO 1", "DESTRANCA, OBRIGATÓRIO"), ("LEGADO 2", "QUALQUER FORMATO")]):
    l = 26 + i*5
    caixa(q, f"B{l}:M{l+1}", " "+rot, AMEIXA, Font(name=TITULO, size=8, bold=True, color=OSSO), ESQ)
    caixa(q, f"O{l}:AY{l+1}", "", PAINEL2, corpo(8, OSSO), ESQ, borda_=AMEIXA)
    caixa(q, f"B{l+2}:AY{l+3}", "  " + fmt, PAINEL, Font(name=CORPO, size=5, color=TEXTO_FRACO), ESQ)
for rot, l, alt in [("LAÇOS   ·   QUEM VOCÊ CONHECE, QUEM TE DEVE, QUEM TE COBRA", 37, 5),
                    ("A INSTITUIÇÃO   ·   O QUE ELA SABE SOBRE VOCÊ", 45, 4),
                    ("PACTO   ·   OPCIONAL, SÓ COM APROVAÇÃO DO MESTRE E PREÇO ESCRITO", 52, 3)]:
    caixa(q, f"B{l}:AY{l+1}", rot, None, rotulo(AMEIXA_CLR, 5))
    caixa(q, f"B{l+2}:AY{l+alt}", "", PAINEL, corpo(8, OSSO), Alignment("left","top",wrap_text=True,indent=1), borda_=AMEIXA_ESC)

faixa(q, "B58:AY59", "REFERÊNCIA RÁPIDA")
tira = [("O turno", "movimento 9 m + ação padrão + ação bônus + reação"),
        ("Arredondamento", "sempre para o lado que não te favorece; o que você ganha nunca fica abaixo de 1"),
        ("Crítico", "20 natural, e dobra os dados. Não dobra Força, nem dado de Melhoria, nem dano fixo"),
        ("Os dois golpes", "feitiço de Toque = os dados da Classe e nada mais, um por turno. Simples = arma + Força"),
        ("Os dois descansos", "curto devolve 25% do PE máximo em ambiente propício; longo zera o relógio e a exaustão")]
for i, (a, b) in enumerate(tira):
    l = 60 + i*2
    cor = PAINEL if i % 2 == 0 else PAINEL2
    caixa(q, f"B{l}:K{l+1}", " " + a, cor, Font(name=TITULO, size=7, bold=True, color=AMEIXA_CLR), ESQ)
    caixa(q, f"M{l}:AY{l+1}", b, cor, corpo(6, TEXTO), ESQ)

# ============================================================ DADOS (oculta)
d = wb.create_sheet("DADOS"); d.sheet_state = "hidden"
d["A1"], d["B1"], d["C1"], d["D1"] = "caminho", "vida1", "vidaN", "peN"
for i, (nome, c) in enumerate(CAT["caminhos"].items(), start=2):
    d[f"A{i}"], d[f"B{i}"], d[f"C{i}"], d[f"D{i}"] = nome, c["vida_inicial"], c["vida_por_nivel"], c["pe_por_nivel"]
n = len(CAT["caminhos"]) + 1
for nome, col in [("caminho_nome","A"),("caminho_vida1","B"),("caminho_vidaN","C"),("caminho_peN","D")]:
    wb.defined_names.add(DefinedName(nome, attr_text=f"DADOS!${col}$2:${col}${n}"))
for col, chave, nome_r in [("F","trilhas","trilha_nome"), ("G",None,"origem_nome")]:
    itens = list(CAT["trilhas"]) if chave else [r["origem"] for r in CAT["rotas_de_criacao"]]
    for i, x in enumerate(itens, start=2): d[f"{col}{i}"] = x
    wb.defined_names.add(DefinedName(nome_r, attr_text=f"DADOS!${col}$2:${col}${len(itens)+1}"))

for ref, fonte in [("AM16","caminho_nome"), ("AM20","trilha_nome"), ("AM24","origem_nome")]:
    dv = DataValidation(type="list", formula1=f"={fonte}", allow_blank=True)
    dv.error = "Escolha um da lista do manual"; dv.errorTitle = "Fora do catálogo"
    ws.add_data_validation(dv); dv.add(ws[ref])
for i in range(5):
    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=6, allow_blank=True)
    dv.error = "Atributo vai de 0 a 6. Na criação, nenhum passa de 3"; dv.errorTitle = "Fora da escala"
    ws.add_data_validation(dv); dv.add(ws[f"I{16+i*4}"])

wb.save("ficha-projeto-m.xlsx")
print("ficha-projeto-m.xlsx  ·  4 abas, catálogos ligados, dropdowns armados")
