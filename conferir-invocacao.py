# -*- coding: utf-8 -*-
"""Confere o invocacao.json contra o capitulo 16 vendorizado, e a planilha
contra o invocacao.json.

O ponto dele e a licao no 9 do projeto: o JSON guarda os valores e o capitulo
guarda a regra, entao alguem tem de comparar os dois. Nenhum valor esperado
esta escrito aqui -- os dois lados sao lidos de arquivo.
"""
import json, os, re, sys

RAIZ = os.path.dirname(os.path.abspath(__file__))
INV = json.load(open(os.path.join(RAIZ, "invocacao.json"), encoding="utf-8"))
CAP_ARQ = os.path.join(RAIZ, "capitulo-16-invocacoes.md")
MAN_ARQ = os.path.join(RAIZ, "manual.txt")
FICHA = os.path.join(RAIZ, "ficha-invocacao", "ficha-invocacao.xlsx")
GERADOR = os.path.join(RAIZ, "ficha-invocacao", "monta.py")

falhas, puladas, checagens = [], [], 0

def checa(nome, ok, detalhe=""):
    global checagens
    checagens += 1
    print(f"  [{'OK' if ok else 'FALHA'}] {nome}" + (f"  <- {detalhe}" if not ok else ""))
    if not ok:
        falhas.append(nome)

def pula(nome, porque):
    puladas.append(nome)
    print(f"  [PULADA] {nome}  <- {porque}")

if not os.path.exists(CAP_ARQ):
    sys.exit("FALHA: o capitulo-16-invocacoes.md nao esta aqui, e sem ele este "
             "validador nao confere NADA.\n"
             "Regere com:\n  cp <clone do JJK---Project>/sistema/05-material/livro/"
             "manual/60-invocacoes.md capitulo-16-invocacoes.md")
CAP = open(CAP_ARQ, encoding="utf-8").read()

CAP35_ARQ = os.path.join(RAIZ, "capitulo-35-caminhos-e-trilhas.md")
if not os.path.exists(CAP35_ARQ):
    sys.exit("FALHA: o capitulo-35-caminhos-e-trilhas.md nao esta aqui, e a Sintonia "
             "do Evocador mora nele.\n"
             "Regere com:\n  cp <clone do JJK---Project>/sistema/05-material/livro/"
             "manual/35-caminhos-e-trilhas.md capitulo-35-caminhos-e-trilhas.md")
CAP35 = open(CAP35_ARQ, encoding="utf-8").read()

def celulas(linha):
    return [c.strip() for c in linha.strip().strip("|").split("|")]

def limpa(s):
    return s.replace("`", "").replace("*", "").strip()

def tabela_depois_de(titulo, ncols):
    """as linhas de dados da primeira tabela que vem depois do titulo."""
    i = CAP.find(titulo)
    if i < 0:
        return []
    linhas, dentro = [], False
    for lin in CAP[i:].split("\n"):
        if lin.startswith("|") and set(limpa(lin).replace("|", "").strip()) <= set("-: "):
            dentro = True
            continue
        if dentro:
            if not lin.startswith("|"):
                break
            cs = celulas(lin)
            if len(cs) == ncols:
                linhas.append(cs)
    return linhas

# =====================================================================
print("=" * 74)
print("1. O CATALOGO — as entradas e o preco de cada uma, contra o capitulo")
print("=" * 74)
for grupo, chave, ncols in [("### `Traço`", "traco", 3), ("### `Comando`", "comando", 3)]:
    do_cap = {}
    for cs in tabela_depois_de(grupo, ncols):
        nome, pts = limpa(cs[1]), limpa(cs[0])
        if pts.isdigit():
            do_cap[nome] = int(pts)
    checa(f"{chave}: o capitulo publica {len(do_cap)} entradas e o json tem as mesmas",
          len(do_cap) == len(INV[chave]),
          f"capitulo {len(do_cap)}, json {len(INV[chave])}: "
          f"{set(do_cap) ^ set(INV[chave])}")
    for nome, pts in INV[chave].items():
        checa(f"{chave} · {nome} = {pts}", do_cap.get(nome) == pts,
              f"o capitulo diz {do_cap.get(nome)!r}")

print()
print("   e o texto de cada entrada sai do capitulo, e nao foi reescrito aqui")
for grupo, chave in [("### `Traço`", "traco"), ("### `Comando`", "comando")]:
    do_cap = {}
    for cs in tabela_depois_de(grupo, 3):
        nome, pts = limpa(cs[1]), limpa(cs[0])
        if pts.isdigit():
            do_cap[nome] = re.sub(r"\s+", " ", limpa(cs[2])).strip()
    fora = []
    for nome, txt_ in INV[chave + "_texto"].items():
        # o Remoto foi reescrito de proposito: a linha do capitulo aponta para
        # outra secao do livro, e um ponteiro nao serve dentro de uma planilha
        if nome == "Remoto":
            continue
        if do_cap.get(nome) != txt_:
            fora.append(nome)
    checa(f"{chave}: todo texto e o do capitulo, palavra por palavra", not fora,
          f"diferem do capitulo: {fora}")
    checa(f"{chave}: o Remoto e a unica reescrita, e ela esta declarada",
          "Remoto" in INV[chave + "_texto"] or chave == "comando")

total = len(INV["traco"]) + len(INV["comando"]) - 1
m = re.search(r"São (\d+) entradas compráveis", CAP)
checa("o capitulo declara o total de compraveis, e ele bate",
      m is not None and int(m.group(1)) == total,
      f"o capitulo diz {m.group(1) if m else '?'}, contei {total}")

print()
print("   toda entrada cai no degrau que a propria regua manda")
for chave, sec in [("traco", "**Régua de `Traço`**"), ("comando", "**Régua de `Comando`**")]:
    degraus = set()
    for cs in tabela_depois_de(sec, 2):
        p = limpa(cs[0])
        if p.isdigit():
            degraus.add(int(p))
    checa(f"{chave}: a regua publica os degraus {sorted(degraus)}",
          degraus == set(INV[chave].values()),
          f"regua {sorted(degraus)}, precos usados {sorted(set(INV[chave].values()))}")
    checa(f"{chave}: os degraus do json sao os mesmos da regua do json",
          {int(k) for k in INV["regua_" + chave]} == degraus,
          f'json {sorted(int(k) for k in INV["regua_" + chave])}, capitulo {sorted(degraus)}')

# =====================================================================
print()
print("=" * 74)
print("2. TIPOS E VIDA — a base de cada tipo e as duas tabelas de nivel")
print("=" * 74)
for cs in tabela_depois_de("**Tipos e vida**", 6):
    nomes = [limpa(x) for x in cs[0].split("·")]
    base = limpa(cs[1])
    if not base.isdigit():
        continue
    for n in nomes:
        checa(f"base do tipo {n} = {base}", INV["tipos"].get(n) == int(base),
              f'o json diz {INV["tipos"].get(n)!r}')
    for j, nivel in enumerate(["2", "10", "18", "30"]):
        esperado = int(limpa(cs[2 + j]))
        calc = int(base) + 2 * int(nivel)
        checa(f"vida crua · {nomes[0]} · nv{nivel} = {esperado}", calc == esperado,
              f"a formula do json da {calc}")

print()
mult = INV["vida"]["multiplicador_corpo_forte"]
checa(f"o capitulo escreve o multiplicador {mult} do corpo forte",
      str(mult).replace(".", ",") in CAP, f"nao achei {mult} no capitulo")
checa("o capitulo pos a Constituicao FORA do multiplicador",
      "+ a Constituição dela × o seu nível" in CAP and INV["vida"]["constituicao_fora_do_multiplicador"])
for cs in tabela_depois_de("**Corpo forte, com Constituição `0`**", 5):
    nomes = [limpa(x) for x in cs[0].split("·")]
    if nomes[0] not in INV["tipos"]:
        continue
    base = INV["tipos"][nomes[0]]
    for j, nivel in enumerate(["2", "10", "18", "30"]):
        esperado = int(limpa(cs[1 + j]))
        calc = int(mult * (base + 2 * int(nivel)))
        checa(f"corpo forte · {nomes[0]} · nv{nivel} = {esperado}", calc == esperado,
              f"a formula do json da {calc}")

# =====================================================================
print()
print("=" * 74)
print("3. O ORCAMENTO — a base, o passo, e a tabela de niveis")
print("=" * 74)
O = INV["orcamento"]
m = re.search(r"O orçamento é `?(\d+)`? no nível 2, e cada marco dá `?\+(\d+)`?", CAP)
checa("o capitulo escreve a base e o passo", m is not None,
      "nao achei a frase do orcamento")
if m:
    checa(f'base = {O["base"]}', int(m.group(1)) == O["base"], f"o capitulo diz {m.group(1)}")
    checa(f'por marco = {O["por_marco"]}', int(m.group(2)) == O["por_marco"],
          f"o capitulo diz {m.group(2)}")
for cs in tabela_depois_de("**Orçamento**", 4):
    nivel, marcos, orc = limpa(cs[0]), limpa(cs[1]), limpa(cs[2])
    if not nivel.isdigit():
        continue
    calc = O["base"] + O["por_marco"] * int(marcos)
    checa(f"orcamento do nv{nivel} = {orc}", calc == int(orc),
          f"a formula do json da {calc}")
    conta_marcos = sum(1 for x in INV["progressao"]["marcos"] if x <= int(nivel))
    checa(f"nv{nivel}: os marcos do json dao {marcos}", conta_marcos == int(marcos),
          f"contei {conta_marcos}")
    checa(f"nv{nivel}: a tabela conferida do json bate",
          O["conferido"].get(nivel) == int(orc), f'json {O["conferido"].get(nivel)!r}')

s_mult = INV["trilhas"]["Servo"]["orcamento_multiplicador"]
m = re.search(r"São `?(\d+)`? no nível 2, `?(\d+)`? no 10, `?(\d+)`? no 18 e `?(\d+)`? no 30", CAP)
checa("o capitulo publica os quatro orcamentos do Servo", m is not None)
if m:
    for nivel, esperado in zip(["2", "10", "18", "30"], m.groups()):
        calc = int((O["base"] + O["por_marco"] *
                    sum(1 for x in INV["progressao"]["marcos"] if x <= int(nivel))) * s_mult)
        checa(f"Servo no nv{nivel} = {esperado}", calc == int(esperado),
              f"o json da {calc}")

# =====================================================================
print()
print("=" * 74)
print("4. O INVESTIR — as sete faixas, nas duas colunas")
print("=" * 74)
do_cap = []
for cs in tabela_depois_de("**Dano do `Investir`**", 3):
    faixa = limpa(cs[0]).replace("–", "-")
    if "-" not in faixa:
        continue
    de, ate = faixa.split("-")
    do_cap.append({"de": int(de), "ate": int(ate), "uma": limpa(cs[1]),
                   "matilha": limpa(cs[2])})
checa(f'o capitulo publica {len(INV["investir"]["faixas"])} faixas',
      len(do_cap) == len(INV["investir"]["faixas"]),
      f'capitulo {len(do_cap)}, json {len(INV["investir"]["faixas"])}')
for a, b in zip(INV["investir"]["faixas"], do_cap):
    checa(f'faixa {b["de"]}-{b["ate"]}: uma = {b["uma"]} · matilha = {b["matilha"]}',
          a == b, f"o json diz {a}")

print()
print("   as faixas cobrem 2 a 30 sem vao nem sobreposicao")
f = INV["investir"]["faixas"]
checa("comeca no 2 e acaba no 30", f[0]["de"] == 2 and f[-1]["ate"] == 30)
checa("cada faixa comeca onde a anterior acabou + 1",
      all(f[i + 1]["de"] == f[i]["ate"] + 1 for i in range(len(f) - 1)))

# =====================================================================
print()
print("=" * 74)
print("5. A FICHA DELA — as linhas que encaram dado, a amarra, a morte")
print("=" * 74)
F, M = INV["ficha_dela"], INV["morte"]
checa("o capitulo diz que os cinco atributos sao DELA",
      "os cinco atributos, e eles são dela" in CAP)
checa("o capitulo diz que ela NAO copia os numeros do dono",
      "não copia os seus números" in CAP)
checa("acerto = o atributo dela + a maestria do dono",
      "o atributo dela + a sua maestria" in CAP)
checa("Defesa = 10 + a Destreza dela + metade da Essencia ou Inteligencia",
      "10 + a Destreza dela + metade da sua Essência" in CAP)
checa(f'ela treina {F["trs_treinados"]} Teste de Resistencia',
      "Ela treina um Teste de Resistência" in CAP and F["trs_treinados"] == 1)
checa(f'deslocamento = {F["deslocamento"]} metros',
      f'| `{F["deslocamento"]}` metros |' in CAP)
checa(f'a amarra = {F["amarra"]} metros',
      f'ficar a até {F["amarra"]} metros' in CAP)
checa("fora da amarra ela NAO some", "Ela **não some**" in CAP or "Ela não some" in CAP)
checa(f'a regua da morte = {M["multiplicador_regua"]} x a vida do TIPO',
      f'A régua da morte é `{M["multiplicador_regua"]} ×` a vida que a fórmula do tipo dá' in CAP)
checa("a regua NAO e a vida daquele corpo",
      "A régua não é a vida daquele corpo" in CAP)
E = M["exemplo_do_capitulo"]
base = INV["tipos"][E["tipo"]]
vmax = base + (2 + E["constituicao"]) * E["nivel"]
checa(f'o exemplo do capitulo fecha: vida {E["vida_maxima"]}', vmax == E["vida_maxima"],
      f"a formula da {vmax}")
checa(f'o exemplo do capitulo fecha: regua {E["regua"]}',
      vmax * M["multiplicador_regua"] == E["regua"],
      f'a formula da {vmax * M["multiplicador_regua"]}')
checa(f'o exemplo do capitulo fecha: volta com {E["volta_com"]}',
      vmax // 2 == E["volta_com"], f"a formula da {vmax // 2}")
checa("o capitulo publica o exemplo com os mesmos numeros",
      f'= {E["vida_maxima"]}` e a régua da morte é `{M["multiplicador_regua"]} × '
      f'{E["vida_maxima"]} = {E["regua"]}' in CAP)

# =====================================================================
print()
print("=" * 74)
print("6. AS TRILHAS — corpos, corpo, multiplicador e area")
print("=" * 74)
for cs in tabela_depois_de("**A vida que cada Trilha põe em campo**", 4):
    nome, corpos = limpa(cs[0]), limpa(cs[1])
    if nome not in INV["trilhas"]:
        continue
    t = INV["trilhas"][nome]
    checa(f"{nome}: {corpos} corpo(s)", t["corpos"] == int(corpos),
          f'o json diz {t["corpos"]}')
    esperado = "cru" if "cru" in cs[2] else "forte"
    checa(f"{nome}: corpo {esperado}", t["corpo"] == esperado, f'o json diz {t["corpo"]!r}')
checa("a Matilha leva metade a mais de area",
      "leva `×1,5` do dano" in CAP and INV["trilhas"]["Matilha"]["vulneravel_a_area"] == 1.5)
checa("area bate uma vez na barra, e nao uma por corpo",
      "não uma vez por corpo" in CAP)

print()
print("   as tres montagens por Trilha gastam o orcamento do nivel 2")
for cs in tabela_depois_de("**Uma montagem pronta por Trilha**", 4):
    nome, entradas, pts = limpa(cs[0]), limpa(cs[1]), limpa(cs[2])
    if nome not in INV["montagens_por_trilha"]:
        continue
    m_ = INV["montagens_por_trilha"][nome]
    checa(f"{nome}: {pts} pontos", m_["pontos"] == int(pts), f'o json diz {m_["pontos"]}')
    soma = sum(INV["traco"].get(e, INV["comando"].get(e, 0)) for e in m_["entradas"])
    checa(f"{nome}: as entradas do json somam {pts}", soma == int(pts), f"somei {soma}")
    orc = int((INV["orcamento"]["base"]) * INV["trilhas"][nome]["orcamento_multiplicador"])
    checa(f"{nome}: gasta o orcamento inteiro do nv2 ({orc})", soma == orc,
          f"gasto {soma}, orcamento {orc}")
    checa(f"{nome}: o arranjo soma {INV['atributos']['pontos_na_criacao']}",
          sum(m_["arranjo"]) == INV["atributos"]["pontos_na_criacao"],
          f'somei {sum(m_["arranjo"])}')
    checa(f'{nome}: nenhum atributo passa do teto da criacao',
          max(m_["arranjo"]) <= INV["atributos"]["teto_na_criacao"],
          f'o maior e {max(m_["arranjo"])}')

print()
print("   as seis montagens do material")
for cs in tabela_depois_de("**Montagens de exemplo**", 5):
    nome, pts, nivel = limpa(cs[0]), limpa(cs[2]), limpa(cs[3])
    achado = [m_ for m_ in INV["montagens_do_material"] if m_["nome"] == nome]
    if not achado:
        continue
    m_ = achado[0]
    checa(f"{nome}: {pts} pontos e cabe no nv{nivel}",
          m_["pontos"] == int(pts) and m_["nivel"] == int(nivel),
          f'o json diz {m_["pontos"]} pontos, nv{m_["nivel"]}')
    soma = sum(INV["traco"].get(e, INV["comando"].get(e, 0)) for e in m_["entradas"])
    checa(f"{nome}: as entradas somam {pts}", soma == int(pts), f"somei {soma}")
    orc = INV["orcamento"]["base"] + INV["orcamento"]["por_marco"] * sum(
        1 for x in INV["progressao"]["marcos"] if x <= m_["nivel"])
    checa(f"{nome}: {pts} pontos cabem no orcamento do nv{nivel} ({orc})", soma <= orc,
          f"gasto {soma}, orcamento {orc}")

# =====================================================================
print()
print("=" * 74)
print("7. O QUE O ORCAMENTO NAO COMPRA")
print("=" * 74)
for item in INV["nao_compra"]:
    checa(f"o capitulo proibe: {item}", item.split(",")[0].lower() in CAP.lower(),
          "nao achei no capitulo")
checa("o capitulo diz que nada do catalogo da Defesa, acerto ou vida direto",
      "nada do catálogo pode dar Defesa, acerto ou vida direto" in CAP)

# =====================================================================
print()
print("=" * 74)
print("8. A PLANILHA — ela sai do json, e nao de numero escrito no gerador")
print("=" * 74)
if not os.path.exists(FICHA):
    pula("a planilha", "ela nao existe. Rode: python3 ficha-invocacao/monta.py")
    pula("o gerador nao guarda valor", "sem a planilha nao ha o que conferir")
else:
    try:
        from openpyxl import load_workbook
    except ImportError:
        pula("a planilha", "o openpyxl nao esta instalado")
    else:
        wb = load_workbook(FICHA)
        checa("as tres abas existem",
              set(wb.sheetnames) == {"INVOCAÇÃO", "CATÁLOGO", "DADOS"},
              str(wb.sheetnames))
        checa("a INVOCACAO abre primeiro", wb.sheetnames[0] == "INVOCAÇÃO",
              wb.sheetnames[0])
        checa("o CATALOGO vem depois dela, e fica VISIVEL",
              wb.sheetnames[1] == "CATÁLOGO"
              and wb["CATÁLOGO"].sheet_state == "visible",
              f'{wb.sheetnames[1]} / {wb["CATÁLOGO"].sheet_state}')
        checa("a DADOS fica escondida", wb["DADOS"].sheet_state == "hidden",
              wb["DADOS"].sheet_state)
        # o catalogo existe para quem RECEBE a planilha: sem o texto, escolher
        # no menu e escolher no escuro.
        _cat = {str(c.value) for lin in wb["CATÁLOGO"].iter_rows() for c in lin
                if c.value is not None}
        _faltam = [n for n in list(INV["traco"]) + list(INV["comando"])
                   if n not in _cat]
        checa("as 20 entradas aparecem na aba CATALOGO", not _faltam,
              f"faltou: {_faltam}")
        _sem_texto = [n for n in list(INV["traco"]) + list(INV["comando"])
                      if INV["traco_texto"].get(n, INV["comando_texto"].get(n, "")) not in _cat]
        checa("cada entrada leva junto o que ela FAZ", not _sem_texto,
              f"sem texto na aba: {_sem_texto}")
        d = wb["DADOS"]
        idx = next((c for c in range(1, 80) if d.cell(row=2, column=c).value == "campo"),
                   None)
        checa("a DADOS publica o indice de celulas", idx is not None)
        if idx:
            n = sum(1 for r in range(3, 200) if d.cell(row=r, column=idx).value)
            campos = {d.cell(row=r, column=idx).value for r in range(3, 200)
                      if d.cell(row=r, column=idx).value}
            # quantos slots a ficha PRECISA ter: derivado do catalogo e do maior
            # orcamento, e nao um numero escrito aqui. Com 4 por grupo a ficha
            # estourava a partir do nivel 6 do Servo.
            def _cabem(precos, orc):
                nn = soma = 0
                for pr in sorted(precos):
                    if soma + pr > orc:
                        break
                    soma += pr
                    nn += 1
                return nn
            _teto = int((INV["orcamento"]["base"] + INV["orcamento"]["por_marco"] *
                         len(INV["progressao"]["marcos"])) *
                        max(t["orcamento_multiplicador"]
                            for t in INV["trilhas"].values()))
            _pt = _cabem(INV["traco"].values(), _teto)
            _pc = _cabem([v for v in INV["comando"].values() if v > 0], _teto)
            tem_t = len([c for c in campos if c.startswith("traco_")])
            tem_c = len([c for c in campos if c.startswith("comando_")])
            checa(f"a ficha tem os {_pt} slots de Traco que o teto pede",
                  tem_t == _pt, f"tem {tem_t}")
            checa(f"a ficha tem os {_pc} slots de Comando que o teto pede",
                  tem_c == _pc, f"tem {tem_c}")
            checa("o indice cobre os slots e os 5 atributos dela",
                  n >= _pt + _pc + 5, f"o indice tem {n} campos")
        # todo Traco e Comando do json esta na aba de apoio
        na_aba = {c.value for lin in d.iter_rows() for c in lin if isinstance(c.value, str)}
        faltando = [e for e in list(INV["traco"]) + list(INV["comando"]) if e not in na_aba]
        checa("todo Traco e Comando do json entrou na planilha", not faltando,
              f"faltou: {faltando}")

G = open(GERADOR, encoding="utf-8").read()
# a regra: o gerador le do json. Um preco do catalogo escrito no codigo e o
# defeito que este validador existe para pegar.
escritos = []
for chave in ("traco", "comando"):
    for nome, pts in INV[chave].items():
        if re.search(rf'"{re.escape(nome)}"\s*:\s*{pts}', G):
            escritos.append(nome)
checa("nenhum preco do catalogo esta escrito dentro do gerador", not escritos,
      f"achei no codigo: {escritos}")

# =====================================================================
print()
print("=" * 74)
print("8b. A SINTONIA — as tres rotas do Evocador, contra o capitulo 35")
print("=" * 74)
S = INV["sintonia"]
checa("o capitulo 35 tem a secao dos degraus do Evocador",
      "### Degraus do Evocador" in CAP35)
checa("o json aponta o capitulo 35 como dono", "capitulo 35" in S["dono"])

# as rotas saem do CAPITULO, e nao do json: o nome de uma checagem nunca pode
# ser contado do lado que ela testa. Este defeito apareceu TRES vezes neste
# validador -- na contagem do Traco, na do Comando e aqui -- e nas tres o
# sintoma era o mesmo: tirar uma entrada do json mudava o nome da checagem, e
# o arnes nao tinha como casar com ela.
_i = CAP35.find("Nível 2: `Sintonia`")
_rotas_cap = set()
if _i >= 0:
    for _lin in CAP35[_i:].split("\n")[1:]:
        if not _lin.strip().startswith(">"):
            break
        _m = re.match(r">\s*\*\*`([^`]+)`\*\*", _lin.strip())
        if _m:
            _rotas_cap.add(_m.group(1))
checa(f"o capitulo 35 publica {len(_rotas_cap)} rotas de Sintonia: "
      f"{sorted(_rotas_cap)}", len(_rotas_cap) == 3, str(sorted(_rotas_cap)))
checa(f"o json tem as mesmas {len(_rotas_cap)} rotas que o capitulo 35",
      set(S["rotas"]) == _rotas_cap,
      f"json {sorted(S['rotas'])}, capitulo {sorted(_rotas_cap)}")
for nome, achar in [("Presa", f'crítico com **{S["rotas"]["Presa"]["critico_a_partir_de"]} ou 20**'),
                    ("Parrudo", f'**`{S["rotas"]["Parrudo"]["multiplicador_maestria"]} ×` a sua maestria**'),
                    ("Voz", f'sobe em **`{S["rotas"]["Voz"]["bonus"]}`**')]:
    checa(f"{nome}: o numero do json e o do capitulo 35", achar in CAP35,
          f"nao achei {achar!r} no capitulo 35")
checa(f'Voz: vira metade da maestria no nivel {S["rotas"]["Voz"]["vira_metade_da_maestria_no_nivel"]}',
      f'a partir do nível {S["rotas"]["Voz"]["vira_metade_da_maestria_no_nivel"]}' in CAP35)
checa("o Parrudo chamava-se Casco, e o json registra isso",
      "Casco" in S["rotas"]["Parrudo"]["chamava_se"])
checa("o capitulo 35 NAO chama mais aquela rota de Casco",
      "**`Casco`**" not in CAP35)

print()
print("   a Voz aponta para um numero que o sistema nao produz — e a ficha")
print("   marca como pendente em vez de chutar")
checa("o capitulo 16 nao escreve NENHUMA formula de CD para a invocacao",
      " CD " not in CAP and "CD da" not in CAP and "CD dela" not in CAP,
      "achei CD no capitulo 16: se ela ganhou formula, a pendencia da Voz fechou")
checa("o json registra a pendencia da Voz", "PENDENTE" in S["rotas"]["Voz"])
checa("a nota que vai PARA A FICHA existe e esta em portugues de gente",
      "não tem fórmula" in S["rotas"]["Voz"].get("nota_na_ficha", "")
      and "combine o número com o mestre" in S["rotas"]["Voz"]["nota_na_ficha"])
checa("a pendencia diz por que ela existe",
      "nao tem formula de cd" in S["rotas"]["Voz"]["PENDENTE"].lower())
checa("o bloco antigo do Parrudo virou ponteiro, e nao segundo dono",
      set(INV["parrudo"]) == {"ponteiro"}, str(list(INV["parrudo"])))

# =====================================================================
print()
print("=" * 74)
print("9. O GUARDA DA DIVERGENCIA — por que este validador nao le o manual.txt")
print("=" * 74)
if not os.path.exists(MAN_ARQ):
    pula("a divergencia com o manual.txt", "o manual.txt nao esta aqui")
else:
    MAN = open(MAN_ARQ, encoding="utf-8", errors="replace").read()
    velho = "A ficha dela é derivada da sua"
    novo = "os cinco atributos, e eles são dela"
    checa("o manual.txt daqui AINDA tem a ficha derivada, que morreu na v0.180",
          velho in MAN,
          "o manual.txt parou de ter a mecanica morta -- se ele foi re-extraido, "
          "esta checagem virou dividida: reveja se o capitulo vendorizado ainda "
          "precisa existir")
    checa("o capitulo vendorizado tem a ficha PROPRIA", novo in CAP)
    checa("os dois discordam MESMO, e e por isso que o dono e o capitulo",
          (velho in MAN) and (velho not in CAP) and (novo in CAP) and (novo not in MAN))
    checa("o json declara por que nao le do manual.txt",
          "congelado" in INV["_meta"]["por_que_nao_le_do_manual_txt"])
    checa("o json declara quem e o dono da regra",
          "capitulo 16" in INV["_meta"]["dono_da_regra"])

# =====================================================================
print()
print("=" * 74)
if falhas:
    print(f">>> {len(falhas)} FALHA(S) de {checagens}")
    for f_ in falhas:
        print(f"    · {f_}")
    sys.exit(1)
if puladas:
    print(f">>> OK, mas {len(puladas)} checagem(ns) PULARAM — e um verde que pulou")
    print("    checagem nao prova nada:")
    for p_ in puladas:
        print(f"    · {p_}")
    sys.exit(0)
print(f">>> TUDO OK — as {checagens} checagens comparam o invocacao.json com o")
print("    capitulo 16, e a planilha com o json. PULADAS: 0")
print("=" * 74)
