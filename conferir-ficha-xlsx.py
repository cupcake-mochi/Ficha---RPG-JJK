# -*- coding: utf-8 -*-
"""Confere o .xlsx gerado contra as decisoes. Nada e digitado aqui.

A checagem que mais vale e a da fonte: o prototipo declarava Oswald e Lexend
e saiu com 4362 celulas em Calibri, porque o openpyxl poe Calibri em toda
celula pintada sem estilo explicito. Um erro que so aparece abrindo o arquivo.
"""
import json, os, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as L

FALHAS = []
def checa(desc, cond, det=""):
    print(f"  [{'OK' if cond else 'FALHA'}] {desc}" + ("" if cond else f"  <- {det}"))
    if not cond: FALHAS.append(desc)

ARQ = "ficha/ficha-projeto-m.xlsx"
if not os.path.exists(ARQ):
    print(f"FALTA O ARQUIVO '{ARQ}'. Gere com:  python3 ficha/monta.py"); sys.exit(1)

CAT = json.load(open("catalogo-projeto-m.json", encoding="utf-8"))
DEC = json.load(open("decisoes-ficha.json", encoding="utf-8"))
wb  = load_workbook(ARQ)

_F = DEC["C4_fontes"]
CORPO = _F["corpo"]["fonte"]
PERMITIDAS = {_F[p]["fonte"] for p in ("corpo", "titulo", "documento", "serie", "marca")}
KANJI_PISO = _F["kanji_piso_pt"]
PALETA = {"120F1D","211C35","30294D","493F54","756588","998BA9","F4F1F7"} | \
         {g["hex"] for g in DEC["A5_acento"]["degraus"]}

print("AS ABAS")
esperadas = DEC["C6_documento"]["abas"]
checa("as cinco abas existem", all(a in wb.sheetnames for a in esperadas),
      str([a for a in esperadas if a not in wb.sheetnames]))
checa("a CARTEIRA abre primeiro (C6: ela é o documento)",
      wb.sheetnames[0] == "CARTEIRA", wb.sheetnames[0])
checa("a DADOS fica escondida", wb["DADOS"].sheet_state == "hidden")

print("\nA FONTE  (o defeito que matou o protótipo)")
contagem, fora = {}, {}
for ws in wb:
    for linha in ws.iter_rows():
        for cel in linha:
            if cel.font and cel.font.name:
                contagem[cel.font.name] = contagem.get(cel.font.name, 0) + 1
                if cel.font.name not in PERMITIDAS:
                    fora.setdefault(cel.font.name, []).append(f"{ws.title}!{cel.coordinate}")
for f, n in sorted(contagem.items(), key=lambda x: -x[1]):
    print(f"      {f:16} {n:>6} células")
checa("nenhuma célula saiu numa fonte que não foi escolhida", not fora,
      str({k: (len(v), v[:2]) for k, v in fora.items()}))
checa("a fonte padrão do documento é a de corpo",
      wb._fonts[0].name == CORPO, str(wb._fonts[0].name))

print("\nTODA FONTE USADA EXISTE MESMO NO SHEETS?")
# Esta checagem existe porque eu propus DUAS fontes que nao existem la, uma
# atras da outra. A lista do Google Fonts nao e a lista do Sheets.
CONF = set(_F["fontes_confirmadas_no_sheets"]["existem"])
NAO  = set(_F["fontes_confirmadas_no_sheets"]["NAO_existem"])
for papel in ("corpo", "titulo", "documento", "serie", "marca"):
    f = _F[papel]["fonte"]
    checa(f"{papel:10} usa '{f}', que foi conferida no seletor de fontes",
          f in CONF, f"'{f}' nao esta na lista conferida"
                     + (" (e esta na lista das que NAO existem)" if f in NAO else ""))
usadas = {c.font.name for ws in wb for l in ws.iter_rows() for c in l
          if c.font and c.font.name}
checa("nenhuma célula usa fonte fora da lista conferida",
      usadas <= CONF, str(sorted(usadas - CONF)))

print("\nO KANJI, E O PISO MEDIDO")
kanji = [(ws.title, c.coordinate, c.font.size, c.value)
         for ws in wb for l in ws.iter_rows() for c in l
         if isinstance(c.value, str) and any("\u4e00" <= ch <= "\u9fff" for ch in c.value)]
for aba, coord, pt, v in kanji:
    print(f"      {aba}!{coord}  {v}  {pt} pt")
checa(f"nenhum kanji abaixo de {KANJI_PISO} pt (abaixo disso o traço funde)",
      all(k[2] >= KANJI_PISO for k in kanji), str([k for k in kanji if k[2] < KANJI_PISO]))

print("\nA ARTE  (C5: desenhada por código, nunca baixada)")
import os as _os
pecas = DEC["C5_arte"]["pecas"]
faltando = [p for p in pecas if not _os.path.exists(f"arte/{p}.png")]
checa(f"as {len(pecas)} peças existem", not faltando, str(faltando))
checa("o gerador da arte está junto", _os.path.exists("arte/gera.py"))
usadas = sum(len(ws._images) for ws in wb)
checa("a ficha usa a arte", usadas >= 5, f"{usadas} imagens")

print("\nA REGRA DURA DA MESA  (o app de celular troca fonte que ele não tem)")
so_corpo = {c.font.name for l in wb["MESA"].iter_rows() for c in l
            if c.font and c.font.name}
checa(f"a MESA usa só {CORPO}", so_corpo <= {CORPO}, str(so_corpo))
checa("a MESA tem 12 colunas com largura definida",
      all(L(c) in wb["MESA"].column_dimensions for c in range(1, 13)))
larg = wb["MESA"].column_dimensions["A"].width
checa("a largura de coluna está na faixa medida (3.6 a 4.1)", 3.6 <= larg <= 4.1, str(larg))

print("\nA LARGURA DAS ABAS DE PC")
# as abas de PC saem da decisão C6, e não de uma lista escrita aqui:
# a TÉCNICA saiu da ficha e esta linha envelheceu junto
for aba in [a for a in DEC["C6_documento"]["abas"] if a not in ("MESA", "DADOS")]:
    n = sum(1 for c in range(1, 60) if L(c) in wb[aba].column_dimensions)
    px = n * (larg * 7)
    checa(f"{aba}: {n} colunas ≈ {px:.0f} px, cabe em notebook de 1366",
          1200 <= px <= 1366, f"{px:.0f} px")

print("\nA COR  (nada fora da paleta decidida)")
fora_c = {}
for ws in wb:
    for linha in ws.iter_rows():
        for cel in linha:
            for cor in (cel.font.color if cel.font else None,):
                if cor is not None and cor.rgb and isinstance(cor.rgb, str):
                    h = cor.rgb[-6:].upper()
                    if h not in PALETA and h != "000000":
                        fora_c.setdefault(h, []).append(f"{ws.title}!{cel.coordinate}")
checa("nenhuma cor de texto fora da paleta", not fora_c,
      str({k: len(v) for k, v in fora_c.items()}))

print("\nAS DECISÕES APARECEM NO ARQUIVO")
ws = wb["DADOS"]
textos = [c.value for l in ws.iter_rows() for c in l if isinstance(c.value, str)]
c1 = DEC["C1_evocador"]
checa(f"o menu de Caminhos tem os {len(c1['caminhos_no_menu'])} que ficaram",
      all(c in textos for c in c1["caminhos_no_menu"]))
menu = [c.value for c in ws["A"][3:3+6] if c.value]
checa("o Evocador NÃO está no menu de Caminhos",
      c1["caminho_oculto"] not in menu, str(menu))
checa("o carimbo de versão está na DADOS",
      str(ws["B1"].value) == CAT["_meta"]["versao"], str(ws["B1"].value))

f = wb["FICHA"]
formulas = [c.value for l in f.iter_rows() for c in l
            if isinstance(c.value, str) and c.value.startswith("=")]
checa("a FICHA tem fórmula, e não número digitado", len(formulas) >= 20, str(len(formulas)))
checa("existe o aviso de catálogo desatualizado (A1)",
      any("a atual é a v" in x for x in formulas))
checa("a Defesa soma uma célula de proteção, não uma constante (C3)",
      any(x.startswith("=10+") and "+$" in x for x in formulas),
      str([x for x in formulas if x.startswith("=10+")][:2]))
checa("a proteção sai do refino, e não de um 1 escrito na mão (C3)",
      any("FLOOR(" in x and "/3" in x for x in formulas))
checa("a maestria usa a lista de marcos, não 'a cada 8 níveis'",
      any('COUNTIF' in x and 'm_maestria' not in x for x in formulas) and
      not any("/8" in x for x in formulas))
checa("o estágio de alma é calculado, não marcado",
      any("estágio 4" in x for x in formulas))

print("\nFORMATO DE DATA E NÚMERO  (o Sheets fala inglês)")
# 'aaaa' nao quebra a formula: ela roda e devolve o dia da semana. Erro que
# so aparece olhando a planilha pronta, entao ele vira checagem.
TOKENS_PT = ["aaaa", "aa/", "/aa", "dd.mm.aaaa"]
todas = [(ws.title, c.coordinate, c.value) for ws in wb for l in ws.iter_rows()
         for c in l if isinstance(c.value, str) and c.value.startswith("=")]
ruins = [(a, co, v) for a, co, v in todas
         if "TEXT(" in v and any(t in v for t in TOKENS_PT)]
checa("nenhum formato de data em português dentro de TEXT()", not ruins,
      str(ruins[:2]))
datas = [v for _, _, v in todas if "TODAY()" in v]
checa("as datas usam yyyy", all("yyyy" in v for v in datas) if datas else True,
      str(datas))

print("\nCADA FÓRMULA PUXA O ATRIBUTO CERTO")
# a ordem do catalogo e Forca, Destreza, Constituicao. Ler atributo por POSICAO
# fazia a Vida somar Destreza; esta checagem existe por causa desse bug.
IDX = {}
dd = wb["DADOS"]
for rr in range(5, 200):
    k, v = dd.cell(row=rr, column=53).value, dd.cell(row=rr, column=54).value
    if k and v: IDX[k] = v
checa("a ficha publica o índice das próprias células", len(IDX) >= 20, str(len(IDX)))
for campo, atributo in [("vida_max", "Constituição"), ("defesa", "Destreza"),
                        ("corpo a corpo", "Força"), ("à distância", "Destreza")]:
    alvo = IDX.get("atr_" + atributo, "?")
    form = f[IDX[campo]].value if campo in IDX else ""
    checa(f"{campo} usa {atributo} ({alvo})",
          isinstance(form, str) and alvo in form.replace("$", ""),
          str(form)[:80])

print("\nO SCRIPT QUE CONSTRÓI A PLANILHA")
import os as _o
GS = "apps-script/Ficha.gs"
checa("o script foi emitido", _o.path.exists(GS))
if _o.path.exists(GS):
    g = open(GS, encoding="utf-8").read()
    checa("ele avisa que é gerado, e não editado na mão", "não edite este arquivo" in g)
    checa("ele traz as seis abas", all(f'"{a}"' in g for a in DEC["C6_documento"]["abas"]))
    checa("ele traz a arte embutida, sem depender de URL",
          '"selo-封.png"' in g and "http" not in g.split("var ARTE")[1][:200])
    checa("ele define as caixas de seleção pela posição medida",
          '"caixas"' in g and "insertCheckboxes" in g)
    checa("a altura de linha vai em pixel, e não em ponto convertido",
          "setRowHeight" in g)
    tam = len(g) / 1024
    checa(f"o arquivo cabe no Apps Script ({tam:.0f} KB, o limite é ~1 MB)", tam < 900)
    # A caixa desmarcada vale FALSO, e FALSO nao e "". A formula antiga somava
    # maestria em TODA pericia com a caixa vazia. A checagem precisa ser
    # exata: so as celulas que REALMENTE tem caixa, lidas do proprio script.
    import re as _re, json as _j
    cx = []
    for bloco in _re.findall(r'"caixas":(\[\[.*?\]\])', g):
        cx += _j.loads(bloco)
    tem_caixa = {f"${L(c)}${r}" for c, r0, n in cx for r in range(r0, r0 + n)}
    checa(f"o script sabe de {len(tem_caixa)} células com caixa de seleção",
          len(tem_caixa) >= 30, str(len(tem_caixa)))
    erradas = [c.value for l in f.iter_rows() for c in l
               if isinstance(c.value, str) and c.value.startswith("=")
               and any(cel + '=""' in c.value for cel in tem_caixa)]
    checa("nenhuma fórmula testa vazio numa célula que tem caixa de seleção",
          not erradas, str(erradas[:2]))

print("\nOS DADOS DO SCRIPT, CONFERIDOS SEM EXECUTAR")
# Nao existe runtime de JavaScript aqui: quem executa Apps Script e o Google.
# O que da para conferir e a METADE de dados do script -- e o 'Range not found'
# que quebrou a montagem era dessa metade, entao vale.
if _o.path.exists(GS):
    import json as _js
    dados = _js.loads(_re.search(r"var ABAS = (\[.*?\]);\n", g, _re.S).group(1))
    nomes = {a["nome"] for a in dados}
    checa("as abas do script são as decididas", nomes == set(DEC["C6_documento"]["abas"]),
          str(nomes ^ set(DEC["C6_documento"]["abas"])))
    fora, artefalta, estilo_ruim = [], [], []
    for a in dados:
        nc, nr = a["cols"], a["rows"]
        for v in a["vals"]:
            if not (1 <= v[0] <= nr and 1 <= v[1] <= nc): fora.append((a["nome"], v[:2]))
            if len(v) > 3 and v[3] >= len(a["estilos"]): estilo_ruim.append((a["nome"], v[:2]))
        for m in a["merges"]:
            if not (m[2] <= nr and m[3] <= nc): fora.append((a["nome"], "merge", m))
        for im in a["imgs"]:
            if not (1 <= im[0] <= nr and 1 <= im[1] <= nc): fora.append((a["nome"], "img", im))
        for cx in a.get("caixas", []):
            if not (1 <= cx[1] and cx[1] + cx[2] - 1 <= nr): fora.append((a["nome"], "caixa", cx))
    checa("nada cai fora dos limites da aba", not fora, str(fora[:3]))
    checa("toda célula com estilo aponta para um estilo que existe", not estilo_ruim,
          str(estilo_ruim[:3]))

    # ESTA e a checagem do erro que quebrou a montagem: menu suspenso aponta
    # para outra aba, e a aba de destino nasce depois. Por isso ele foi movido
    # para o fim -- e por isso a checagem confere que a origem existe.
    ruins = []
    for a in dados:
        for alvo, fonte in a["dv"]:
            aba_fonte = fonte.split("!")[0].strip("=") if "!" in fonte else a["nome"]
            if aba_fonte not in nomes:
                ruins.append((a["nome"], fonte))
    checa("todo menu suspenso puxa de uma aba que existe", not ruins, str(ruins[:3]))
    checa("os menus são aplicados só depois de todas as abas nascerem",
          "menusSuspensos_" in g and "setDataValidation" not in
          g[g.index("function montarAba_"):g.index("function mat_")])

    arte_usada = {im[4] for a in dados for im in a["imgs"]}
    embutida = set(_re.findall(r'"([^"]+\.png)":"', g.split("var ARTE")[1][:200000]))
    checa("toda imagem usada está embutida no script", arte_usada <= embutida,
          str(sorted(arte_usada - embutida)))
    fontes_gs = {e[0] for a in dados for e in a["estilos"]}
    checa("o script só usa fonte conferida no seletor", fontes_gs <= CONF,
          str(sorted(fontes_gs - CONF)))

print(f"\n{'A FICHA CONFERE' if not FALHAS else f'{len(FALHAS)} PROBLEMA(S)'}")
sys.exit(1 if FALHAS else 0)
