# -*- coding: utf-8 -*-
"""Monta a ficha da invocacao como arquivo SOLTO -- o que se manda para alguem.

O desenho todo mora no constroi.py, porque ele tem dois destinos: este arquivo
e a aba da ficha do personagem (ficha-v01/monta.py). Aqui as tres celulas do
dono sao digitadas; como aba, elas puxam da FICHA.
"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(AQUI), "ficha"))
sys.path.insert(0, AQUI)
from estilo import CORPO, PT_VALOR, TEXTO
from constroi import constroi

wb = Workbook()
wb.remove(wb.active)
wb._fonts[0] = Font(name=CORPO, size=PT_VALOR, color=TEXTO)

info = constroi(wb)

for i, aba in enumerate(["INVOCAÇÃO", "CATÁLOGO"]):
    wb.move_sheet(aba, i - wb.sheetnames.index(aba))

saida = os.path.join(AQUI, "ficha-invocacao.xlsx")
wb.save(saida)
print(f"ficha escrita: {saida}")
print(f"abas: {wb.sheetnames}")
print(f"campos indexados: {len(info['campos'])} · "
      f"slots: {info['n_traco']} Traço e {info['n_comando']} Comando")
