# -*- coding: utf-8 -*-
"""Renderiza a ficha da invocacao em HTML, lendo o .xlsx.

Ele nao redesenha nada: le a largura de cada coluna, a altura de cada linha, as
mesclagens, o preenchimento, a fonte, o tamanho, a cor e o alinhamento de cada
celula, e monta a grade. O que a pagina mostra e a planilha.

As formulas sao recalculadas pelo motor `formulas` com um exemplo preenchido,
para a previa mostrar numero e nao `=IF(...)`.

    python3 ficha-invocacao/previa.py > previa.html
"""
import json, os, sys, warnings
warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as L

AQUI = os.path.dirname(os.path.abspath(__file__))
FICHA = os.path.join(AQUI, "ficha-invocacao.xlsx")
ABAS = ["INVOCAÇÃO", "CATÁLOGO"]

# --- o exemplo, para a previa nao ficar vazia ------------------------------
EXEMPLO = {
    "nivel": 10, "ess_dono": 4, "int_dono": 2, "nome": "Cão de Cinzas",
    "tipo": "técnica", "trilha": "Matilha", "sintonia": "Parrudo",
    "atr_Força": 3, "atr_Destreza": 3, "atr_Constituição": 2,
    "atr_Inteligência": 1, "atr_Essência": 2,
    "atr_acerto": "Destreza", "defesa_de": "Essência",
    "tr_treinado": "Físico", "fisico_de": "Destreza", "vida": 61,
    "traco_1": "Faro", "comando_1": "Agarrar", "comando_2": "Arrastar",
}

wb = load_workbook(FICHA)
d = wb["DADOS"]
IDX = next(c for c in range(1, 80) if d.cell(row=2, column=c).value == "campo")
CEL = {}
for r in range(3, 200):
    k = d.cell(row=r, column=IDX).value
    if not k:
        break
    CEL[k] = d.cell(row=r, column=IDX + 1).value

VALORES = {}
try:
    import formulas
    M = formulas.ExcelModel().loads(FICHA).finish()
    s0 = M.calculate()
    pref = [k for k in s0 if "INVOCAÇÃO'!" in k][0].split("INVOCAÇÃO'!")[0]
    sol = M.calculate(inputs={pref + "INVOCAÇÃO'!" + CEL[k]: v
                              for k, v in EXEMPLO.items() if k in CEL})
    for k, v in sol.items():
        try:
            val = v.value[0, 0]
        except Exception:
            continue
        VALORES[k] = val
    PREF = pref
except Exception as exc:
    print(f"<!-- sem recalculo: {exc} -->", file=sys.stderr)
    PREF = None

def calculado(aba, coord):
    if not PREF:
        return None
    return VALORES.get(f"{PREF}'[{os.path.basename(FICHA).lower()}]{aba}'!{coord}"
                       .replace("''", "'"))

def cor(c, padrao=None):
    if c is None or getattr(c, "rgb", None) in (None, "00000000"):
        return padrao
    rgb = c.rgb
    return "#" + rgb[2:] if isinstance(rgb, str) and len(rgb) == 8 else padrao

# as celulas de ENTRADA guardam o padrao do arquivo, e o motor sobrescreve elas
# so em memoria. Sem isto a previa mostrava vida 85 com nivel 2 na tela.
ENTRADAS = {CEL[k]: v for k, v in EXEMPLO.items() if k in CEL}

def mostra(cel, aba):
    if aba == "INVOCAÇÃO" and cel.coordinate in ENTRADAS:
        v = ENTRADAS[cel.coordinate]
        if isinstance(v, float) and v == int(v):
            v = int(v)
        return str(v)
    v = cel.value
    if isinstance(v, str) and v.startswith("="):
        # o valor recalculado, quando existe
        chave = [k for k in VALORES if k.endswith(f"{aba}'!{cel.coordinate}")]
        if chave:
            v = VALORES[chave[0]]
        else:
            return ""
    if v is None:
        return ""
    if v is False or v is True:
        return "☐" if v is False else "☑"
    if isinstance(v, float) and v == int(v):
        v = int(v)
    return str(v)

FONTES = {"Oswald": "'Oswald', 'Roboto Condensed', sans-serif",
          "Roboto": "'Roboto', system-ui, sans-serif",
          "Castoro": "'Castoro', Georgia, serif",
          "Courier New": "'Courier New', monospace",
          "Yuji Syuku": "'Yuji Syuku', serif"}

PX = 7.2   # largura de coluna -> pixel

def render(nome):
    ws = wb[nome]
    largura = {}
    for k, v in ws.column_dimensions.items():
        if v.width:
            for c in range(v.min, v.max + 1):
                largura[c] = v.width
    padrao_larg = 4.0
    ncol = ws.max_column
    nlin = ws.max_row
    coberta = {}
    ancora = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                coberta[(r, c)] = True
        coberta.pop((rng.min_row, rng.min_col), None)
        ancora[(rng.min_row, rng.min_col)] = (
            rng.max_col - rng.min_col + 1, rng.max_row - rng.min_row + 1)

    out = ['<div class="grade-fora"><table class="grade"><colgroup>']
    for c in range(1, ncol + 1):
        out.append(f'<col style="width:{largura.get(c, padrao_larg) * PX:.1f}px">')
    out.append("</colgroup><tbody>")
    for r in range(1, nlin + 1):
        alt = ws.row_dimensions[r].height or 15.0
        out.append(f'<tr style="height:{alt * 1.34:.1f}px">')
        for c in range(1, ncol + 1):
            if coberta.get((r, c)):
                continue
            cel = ws.cell(row=r, column=c)
            cs, rs = ancora.get((r, c), (1, 1))
            f, p, a = cel.font, cel.fill, cel.alignment
            est = []
            bg = cor(p.start_color) if (p and p.fill_type == "solid") else None
            if bg: est.append(f"background:{bg}")
            if f and f.name: est.append(f"font-family:{FONTES.get(f.name, 'inherit')}")
            if f and f.sz: est.append(f"font-size:{f.sz * 1.05:.1f}px")
            fc = cor(f.color) if f else None
            if fc: est.append(f"color:{fc}")
            if f and f.b: est.append("font-weight:600")
            if a and a.horizontal: est.append(f"text-align:{a.horizontal}")
            if a and a.text_rotation: est.append("writing-mode:vertical-rl")
            txt_ = mostra(cel, nome)
            esc = (txt_.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
            atr = f' colspan="{cs}"' if cs > 1 else ""
            atr += f' rowspan="{rs}"' if rs > 1 else ""
            out.append(f'<td{atr} style="{";".join(est)}">{esc}</td>')
        out.append("</tr>")
    out.append("</tbody></table></div>")
    return "".join(out)

print(json.dumps({n: render(n) for n in ABAS}, ensure_ascii=False))
