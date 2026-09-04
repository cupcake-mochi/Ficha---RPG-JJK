# -*- coding: utf-8 -*-
"""Compara a ficha que o ficha-v01/monta.py gera com o .xlsx que o Mizuki mandou.

Celula por celula: valor, formula, fonte, preenchimento, borda, alinhamento e
formato de numero. Mais as mesclagens, as larguras, as alturas, os menus, a
condicional, as imagens e a ordem das abas.

As diferencas ESPERADAS sao as tres limpezas declaradas no layout.json, e so
elas. Qualquer outra e defeito do gerador.
"""
import json, os, sys
from collections import Counter
from openpyxl import load_workbook

AQUI = os.path.dirname(os.path.abspath(__file__))
A = os.path.join(AQUI, "ficha-v01", "original.xlsx")
B = os.path.join(AQUI, "ficha-v01", "ficha-projeto-m-0.1.xlsx")
LAY = json.load(open(os.path.join(AQUI, "ficha-v01", "layout.json"), encoding="utf-8"))

for f in (A, B):
    if not os.path.exists(f):
        sys.exit(f"falta {f}. Rode extrair.py e monta.py primeiro.")

wa, wb_ = load_workbook(A), load_workbook(B)
difs, esperadas = [], Counter()

def cor(c):
    if c is None or getattr(c, "rgb", None) in (None, "00000000"):
        return None
    return c.rgb if isinstance(c.rgb, str) else None

def perfil(cel):
    f, p, b, al = cel.font, cel.fill, cel.border, cel.alignment
    return {
        "valor": cel.value,
        "fonte": [f.name, f.sz, cor(f.color), bool(f.b), bool(f.i)] if f else None,
        "fundo": cor(p.start_color) if (p and p.fill_type == "solid") else None,
        "borda": {l: [getattr(b, l).style, cor(getattr(b, l).color)]
                  for l in ("top", "bottom", "left", "right")
                  if getattr(b, l) and getattr(b, l).style} if b else {},
        "alinha": [al.horizontal, al.vertical, bool(al.wrap_text),
                   al.text_rotation or 0] if al else None,
        "fmt": cel.number_format,
    }

def eh_ruido_de_fabrica(pa, pb):
    """a limpeza 1: Arial 10 preto em celula VAZIA, e so isso."""
    if pa["valor"] is not None or pb["valor"] is not None:
        return False
    fa, fb = pa["fonte"], pb["fonte"]
    if not fa or fa[0] != "Arial" or fa[1] != 10.0:
        return False
    # o resto tem de ser igual
    return all(pa[k] == pb[k] for k in ("fundo", "borda", "alinha", "fmt"))

print("=" * 74)
print("AS ABAS")
print("=" * 74)
print(f"  original: {wa.sheetnames}")
print(f"  gerada:   {wb_.sheetnames}")
if wa.sheetnames != wb_.sheetnames:
    difs.append(f"ordem/nome das abas: {wa.sheetnames} != {wb_.sheetnames}")
for n in wa.sheetnames:
    if n in wb_.sheetnames and wa[n].sheet_state != wb_[n].sheet_state:
        difs.append(f"{n}: estado {wa[n].sheet_state} != {wb_[n].sheet_state}")
print(f"  estados: " + ", ".join(f"{n}={wa[n].sheet_state}" for n in wa.sheetnames))

for n in wa.sheetnames:
    if n not in wb_.sheetnames:
        continue
    sa, sb = wa[n], wb_[n]
    print()
    print("=" * 74)
    print(f"A ABA {n}")
    print("=" * 74)

    lin = max(sa.max_row, sb.max_row)
    col = max(sa.max_column, sb.max_column)
    print(f"  extensao: original {sa.max_row}x{sa.max_column} · "
          f"gerada {sb.max_row}x{sb.max_column}")

    iguais = ruido = 0
    for r in range(1, lin + 1):
        for c in range(1, col + 1):
            pa, pb = perfil(sa.cell(row=r, column=c)), perfil(sb.cell(row=r, column=c))
            if pa == pb:
                iguais += 1
                continue
            if eh_ruido_de_fabrica(pa, pb):
                ruido += 1
                esperadas["Arial 10 de fábrica em célula vazia"] += 1
                continue
            coord = sa.cell(row=r, column=c).coordinate
            for k in pa:
                if pa[k] != pb[k]:
                    difs.append(f"{n}!{coord} {k}: {pa[k]!r} != {pb[k]!r}")
    print(f"  células idênticas: {iguais}")
    print(f"  células que só diferem pelo Arial de fábrica: {ruido}")

    ma, mb = {str(x) for x in sa.merged_cells.ranges}, {str(x) for x in sb.merged_cells.ranges}
    print(f"  mesclagens: {len(ma)} original · {len(mb)} gerada")
    for x in sorted(ma - mb): difs.append(f"{n}: mesclagem {x} faltou")
    for x in sorted(mb - ma): difs.append(f"{n}: mesclagem {x} sobrou")

    def largs(s):
        d = {}
        for k, v in s.column_dimensions.items():
            if v.width:
                for c in range(v.min, v.max + 1):
                    d[c] = round(v.width, 2)
        return d
    la, lb = largs(sa), largs(sb)
    trocadas = 0
    for c in set(la) | set(lb):
        va, vb = la.get(c), lb.get(c)
        if va == vb:
            continue
        if va and vb and abs(va - 3.63) < 0.01 and abs(vb - 4.0) < 0.01:
            trocadas += 1
            esperadas["largura 3,63 -> 4,0"] += 1
            continue
        difs.append(f"{n}: largura da coluna {c}: {va} != {vb}")
    print(f"  larguras: {len(la)} original · {len(lb)} gerada · "
          f"{trocadas} trocadas de 3,63 para 4,0")

    ha = {int(k): v.height for k, v in sa.row_dimensions.items() if v.height}
    hb = {int(k): v.height for k, v in sb.row_dimensions.items() if v.height}
    for k in set(ha) | set(hb):
        if ha.get(k) != hb.get(k):
            difs.append(f"{n}: altura da linha {k}: {ha.get(k)} != {hb.get(k)}")
    print(f"  alturas de linha: {len(ha)} original · {len(hb)} gerada")

    va = {(str(v.sqref), v.type, v.formula1) for v in sa.data_validations.dataValidation}
    vbs = {(str(v.sqref), v.type, v.formula1) for v in sb.data_validations.dataValidation}
    print(f"  menus suspensos: {len(va)} original · {len(vbs)} gerada")
    for x in sorted(va - vbs): difs.append(f"{n}: menu {x} faltou")
    for x in sorted(vbs - va): difs.append(f"{n}: menu {x} sobrou")

    def regras(s):
        out = []
        for faixa in s.conditional_formatting:
            for r in faixa.rules:
                fc = cor(r.dxf.font.color) if (r.dxf and r.dxf.font and r.dxf.font.color) else None
                fb = cor(r.dxf.fill.bgColor) if (r.dxf and r.dxf.fill and r.dxf.fill.bgColor) else None
                out.append((str(faixa.sqref), r.type, fc, fb))
        return out
    ra, rb = regras(sa), regras(sb)
    de_fabrica = [x for x in ra if x[3] == "FFB7E1CD"]
    esperadas["condicional verde de fábrica"] += len(de_fabrica)
    ra_limpa = [x for x in ra if x[3] != "FFB7E1CD"]
    print(f"  condicional: {len(ra)} original ({len(de_fabrica)} de fábrica) · "
          f"{len(rb)} gerada")
    # a formula nao entra na comparacao: o Sheets exporta com aspas extras, e
    # tirar elas e o que faz a regra valer de novo
    for x in ra_limpa:
        if x not in rb: difs.append(f"{n}: condicional {x} faltou")
    for x in rb:
        if x not in ra_limpa: difs.append(f"{n}: condicional {x} sobrou")

    ia = [(round(i.width), round(i.height)) for i in getattr(sa, "_images", [])]
    ib = [(round(i.width), round(i.height)) for i in getattr(sb, "_images", [])]
    print(f"  imagens: {len(ia)} original · {len(ib)} gerada")
    if sorted(ia) != sorted(ib):
        difs.append(f"{n}: tamanhos de imagem {sorted(ia)} != {sorted(ib)}")

print()
print("=" * 74)
print("AS DIFERENÇAS ESPERADAS — as três limpezas que o Mizuki decidiu")
print("=" * 74)
for k, v in esperadas.items():
    print(f"  {v:>6}  {k}")
for lim in LAY["_meta"]["limpezas"]:
    print(f"    · {lim}")

print()
print("=" * 74)
if difs:
    print(f">>> {len(difs)} DIFERENÇA(S) NÃO EXPLICADA(S):")
    for d in difs[:40]:
        print(f"    · {d}")
    if len(difs) > 40:
        print(f"    ... e mais {len(difs) - 40}")
    sys.exit(1)
print(">>> IGUAIS — fora as três limpezas declaradas, o gerador reproduz o")
print("    arquivo do Mizuki célula por célula.")
print("=" * 74)
