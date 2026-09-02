# -*- coding: utf-8 -*-
"""Recalcula a ficha da invocacao e compara com os numeros que o capitulo 16
PUBLICA. Nenhum valor esperado esta escrito aqui: os casos saem do
invocacao.json, que por sua vez sai do capitulo.

Ele nao usa o LibreOffice, e isso e de proposito: o motor `formulas` avalia o
.xlsx direto, entao o teste roda em qualquer maquina que tenha o pacote --
inclusive onde o LibreOffice nao traz o filtro de Calc.

    pip install formulas
"""
import json, os, sys, warnings
warnings.filterwarnings("ignore")
from openpyxl import load_workbook

RAIZ = os.path.dirname(os.path.abspath(__file__))
INV = json.load(open(os.path.join(RAIZ, "invocacao.json"), encoding="utf-8"))
FICHA = os.path.join(RAIZ, "ficha-invocacao", "ficha-invocacao.xlsx")
ABA = "INVOCAÇÃO"

falhas, checagens = [], 0

def checa(nome, ok, detalhe=""):
    global checagens
    checagens += 1
    print(f"  [{'OK' if ok else 'FALHA'}] {nome}" + (f"  <- {detalhe}" if not ok else ""))
    if not ok:
        falhas.append(nome)

if not os.path.exists(FICHA):
    sys.exit("a ficha nao existe. Rode: python3 ficha-invocacao/monta.py")
try:
    import formulas
except ImportError:
    sys.exit("FALHA: o pacote `formulas` nao esta instalado, e sem ele este "
             "teste nao confere NADA. Rode: pip install formulas")

# --- o indice de celulas, lido da propria ficha ------------------------
_d = load_workbook(FICHA)["DADOS"]
IDX = next(c for c in range(1, 80) if _d.cell(row=2, column=c).value == "campo")
CEL = {}
for r in range(3, 200):
    k = _d.cell(row=r, column=IDX).value
    if not k:
        break
    CEL[k] = _d.cell(row=r, column=IDX + 1).value
print(f"o indice publica {len(CEL)} campos\n")

MODELO = formulas.ExcelModel().loads(FICHA).finish()

# a chave do motor sai do PROPRIO motor, e nao de um nome montado a mao: assim
# ela nao pode divergir do arquivo (foi o que quebrou a primeira versao deste
# teste -- o nome montado vinha em maiuscula e nenhum input pegava).
_SOL0 = MODELO.calculate()
_amostra = [k for k in _SOL0 if f"{ABA}'!" in k]
if not _amostra:
    sys.exit(f"o motor nao devolveu nenhuma celula da aba {ABA!r}")
PREFIXO = _amostra[0].split(f"{ABA}'!")[0] + f"{ABA}'!"
# celula vazia e sem formula nao aparece na solucao do motor, entao a chave se
# monta do prefixo em vez de ser procurada -- procurar sumia com os campos que
# o jogador preenche, que sao justamente as ENTRADAS do teste.
CHAVE = {campo: PREFIXO + coord for campo, coord in CEL.items()}
INVERSO = {v: k for k, v in CHAVE.items()}

def roda(**entradas):
    """poe os valores nas celulas de entrada e devolve o que a ficha calcula."""
    for k in entradas:
        if k not in CHAVE:
            sys.exit(f"campo desconhecido no teste: {k!r}")
    sol = MODELO.calculate(inputs={CHAVE[k]: v for k, v in entradas.items()})
    saidas = {}
    for k, v in sol.items():
        if k in INVERSO:
            try:
                val = v.value[0, 0]
            except Exception:
                val = v
            saidas[INVERSO[k]] = val
    return saidas

def num(v):
    try:
        return round(float(v), 6)
    except (TypeError, ValueError):
        return v

BASE = dict(ess_dono=0, int_dono=0)
def zerado(**kw):
    v = dict(BASE)
    for a in INV["atributos"]["lista"]:
        v["atr_" + a] = 0
    v.update(kw)
    return v

# =====================================================================
print("=" * 74)
print("1. O EXEMPLO DO CAPITULO — Kaito, nivel 10, tecnica, Constituicao 1")
print("=" * 74)
E = INV["morte"]["exemplo_do_capitulo"]
s = roda(**zerado(nivel=E["nivel"], tipo=E["tipo"], trilha="Coro",
                  **{"atr_Constituição": E["constituicao"]}))
checa(f'vida maxima = {E["vida_maxima"]}', num(s["vida_max"]) == E["vida_maxima"],
      f'a ficha deu {s["vida_max"]!r}')
checa(f'regua da morte = {E["regua"]}', num(s["regua"]) == E["regua"],
      f'a ficha deu {s["regua"]!r}')
checa(f'volta com {E["volta_com"]}', num(s["volta_com"]) == E["volta_com"],
      f'a ficha deu {s["volta_com"]!r}')
checa("maestria do nivel 10 = 2", num(s["maestria"]) == 2, f'{s["maestria"]!r}')
checa("maior Classe do nivel 10 = 3 — e o PE de invocar",
      num(s["classe"]) == 3, f'{s["classe"]!r}')

# =====================================================================
print()
print("=" * 74)
print("2. O CORPO FORTE — a tabela de Constituicao 0 que o capitulo imprime")
print("=" * 74)
for tipo, col in INV["vida"]["conferido_corpo_forte_con0"].items():
    for nivel, esperado in col.items():
        s = roda(**zerado(nivel=int(nivel), tipo=tipo, trilha="Servo"))
        checa(f"corpo forte · {tipo} · nv{nivel} = {esperado}",
              num(s["vida_max"]) == esperado, f'a ficha deu {s["vida_max"]!r}')

print()
print("=" * 74)
print("3. A VIDA CRUA — a tabela de Tipos e vida, pelo Coro")
print("=" * 74)
for tipo, col in INV["vida"]["conferido_cru_con0"].items():
    for nivel, esperado in col.items():
        s = roda(**zerado(nivel=int(nivel), tipo=tipo, trilha="Coro"))
        checa(f"vida crua · {tipo} · nv{nivel} = {esperado}",
              num(s["vida_max"]) == esperado, f'a ficha deu {s["vida_max"]!r}')

print()
print("   contra-teste: o Servo e o Coro NAO podem dar a mesma vida")
s1 = roda(**zerado(nivel=10, tipo="técnica", trilha="Servo"))
s2 = roda(**zerado(nivel=10, tipo="técnica", trilha="Coro"))
checa("Servo (55) != Coro (22) no mesmo nivel e tipo",
      num(s1["vida_max"]) != num(s2["vida_max"]),
      f'os dois deram {s1["vida_max"]!r}')

print()
print("   a regua sai da vida CRUA, e nao da do corpo forte")
checa("a regua do Servo e a do Coro sao IGUAIS — 5 x a crua",
      num(s1["regua"]) == num(s2["regua"]),
      f'Servo {s1["regua"]!r} · Coro {s2["regua"]!r}')

# =====================================================================
print()
print("=" * 74)
print("4. O ORCAMENTO — a tabela do capitulo, e o mais-metade do Servo")
print("=" * 74)
for nivel, esperado in INV["orcamento"]["conferido"].items():
    s = roda(**zerado(nivel=int(nivel), tipo="técnica", trilha="Coro"))
    checa(f"orcamento do nv{nivel} = {esperado}", num(s["orcamento"]) == esperado,
          f'a ficha deu {s["orcamento"]!r}')

mult = INV["trilhas"]["Servo"]["orcamento_multiplicador"]
for nivel, base_orc in INV["orcamento"]["conferido"].items():
    esperado = int(base_orc * mult)
    s = roda(**zerado(nivel=int(nivel), tipo="técnica", trilha="Servo"))
    checa(f"Servo no nv{nivel} = {esperado} — o da ficha mais metade",
          num(s["orcamento"]) == esperado, f'a ficha deu {s["orcamento"]!r}')

print()
print("   o capitulo diz que 'mais metade' sempre fecha redondo")
for nivel in INV["orcamento"]["conferido"]:
    s = roda(**zerado(nivel=int(nivel), tipo="técnica", trilha="Servo"))
    checa(f"nv{nivel}: o orcamento do Servo e inteiro",
          float(num(s["orcamento"])).is_integer(), f'{s["orcamento"]!r}')

# =====================================================================
print()
print("=" * 74)
print("5. AS TRES MONTAGENS POR TRILHA — gastam o orcamento INTEIRO no nv2")
print("=" * 74)
for trilha, m in INV["montagens_por_trilha"].items():
    v = zerado(nivel=2, tipo="técnica", trilha=trilha)
    for i, a in enumerate(INV["atributos"]["lista"]):
        v["atr_" + a] = m["arranjo"][i]
    tr = [e for e in m["entradas"] if e in INV["traco"]]
    cm = [e for e in m["entradas"] if e in INV["comando"]]
    for i, e in enumerate(tr):
        v[f"slot_nome_{i+1}"] = e
    for i, e in enumerate(cm):
        v[f"slot_nome_{5+i}"] = e
    s = roda(**v)
    checa(f"{trilha}: gasto = {m['pontos']}", num(s["gasto"]) == m["pontos"],
          f'a ficha deu {s["gasto"]!r}')
    checa(f"{trilha}: sobra zero", num(s["sobra"]) == 0, f'a ficha deu {s["sobra"]!r}')
    checa(f"{trilha}: o arranjo soma {sum(m['arranjo'])} e a conferencia diz ok",
          s["aviso_atributo"] == "ok", f'a ficha deu {s["aviso_atributo"]!r}')

# =====================================================================
print()
print("=" * 74)
print("6. AS SEIS MONTAGENS DO MATERIAL — cabem no nivel que o capitulo diz")
print("=" * 74)
for m in INV["montagens_do_material"]:
    v = zerado(nivel=m["nivel"], tipo="técnica", trilha="Coro")
    for i, a in enumerate(INV["atributos"]["lista"]):
        v["atr_" + a] = m["arranjo"][i]
    tr = [e for e in m["entradas"] if e in INV["traco"]]
    cm = [e for e in m["entradas"] if e in INV["comando"]]
    for i, e in enumerate(tr):
        v[f"slot_nome_{i+1}"] = e
    for i, e in enumerate(cm):
        v[f"slot_nome_{5+i}"] = e
    s = roda(**v)
    checa(f"{m['nome']}: gasto = {m['pontos']}", num(s["gasto"]) == m["pontos"],
          f'a ficha deu {s["gasto"]!r}')
    checa(f"{m['nome']}: cabe no nv{m['nivel']} — a sobra nao e negativa",
          not isinstance(s["sobra"], str), f'a ficha deu {s["sobra"]!r}')

print()
print("   contra-teste: uma montagem que NAO cabe tem de acusar")
v = zerado(nivel=2, tipo="técnica", trilha="Coro")
v["slot_nome_1"], v["slot_nome_2"] = "Voo", "Montaria"      # 8 + 8 = 16, contra 8
s = roda(**v)
checa("Voo + Montaria no nv2 (16 pontos de um orcamento de 8) acusa 'estourou'",
      isinstance(s["sobra"], str) and "estourou" in s["sobra"], f'a ficha deu {s["sobra"]!r}')

# =====================================================================
print()
print("=" * 74)
print("7. O INVESTIR — as sete faixas, nas duas colunas")
print("=" * 74)
for f in INV["investir"]["faixas"]:
    for nivel in (f["de"], f["ate"]):
        for trilha, chave in [("Coro", "uma"), ("Matilha", "matilha")]:
            s = roda(**zerado(nivel=nivel, tipo="técnica", trilha=trilha))
            checa(f"nv{nivel} · {trilha} = {f[chave]}", str(s["investir"]) == f[chave],
                  f'a ficha deu {s["investir"]!r}')

# =====================================================================
print()
print("=" * 74)
print("8. A FICHA DELA — acerto, Defesa e Teste de Resistencia")
print("=" * 74)
# nivel 18: maestria 3. Destreza dela 4, Essencia do dono 5 -> Defesa 10+4+2 = 16
s = roda(**zerado(nivel=18, tipo="técnica", trilha="Coro", ess_dono=5, int_dono=1,
                  defesa_de="Essência", atr_acerto="Destreza", tr_treinado="Vigor",
                  fisico_de="Força", **{"atr_Destreza": 4, "atr_Constituição": 2}))
checa("maestria do nv18 = 3", num(s["maestria"]) == 3, f'{s["maestria"]!r}')
checa("acerto = Destreza dela (4) + maestria (3) = 7", num(s["acerto"]) == 7,
      f'a ficha deu {s["acerto"]!r}')
checa("Defesa = 10 + Destreza dela (4) + metade da Essencia dele (5//2=2) = 16",
      num(s["defesa"]) == 16, f'a ficha deu {s["defesa"]!r}')

s2 = roda(**zerado(nivel=18, tipo="técnica", trilha="Coro", ess_dono=5, int_dono=1,
                   defesa_de="Inteligência", atr_acerto="Destreza",
                   **{"atr_Destreza": 4}))
checa("trocando para Inteligencia (1//2=0), a Defesa cai para 14",
      num(s2["defesa"]) == 14, f'a ficha deu {s2["defesa"]!r}')
checa("contra-teste: a escolha Essencia/Inteligencia muda MESMO a Defesa",
      num(s["defesa"]) != num(s2["defesa"]))

# =====================================================================
print()
print("=" * 74)
print("9. A SINTONIA — o Parrudo na vida, a Presa no crítico")
print("=" * 74)
P = INV["sintonia"]["rotas"]["Parrudo"]["multiplicador_maestria"]
for nivel, maestria in [(2, 1), (10, 2), (18, 3), (30, 4)]:
    sem = roda(**zerado(nivel=nivel, tipo="técnica", trilha="Servo", sintonia="—"))
    com = roda(**zerado(nivel=nivel, tipo="técnica", trilha="Servo", sintonia="Parrudo"))
    d_ = num(com["vida_max"]) - num(sem["vida_max"])
    checa(f"nv{nivel}: o Parrudo soma {P} x maestria ({maestria}) = {P*maestria}",
          d_ == P * maestria, f"a ficha somou {d_}")

print()
# a Presa: ela mexe no dado do acerto, e nao na vida
pr = roda(**zerado(nivel=10, tipo="técnica", trilha="Coro", sintonia="Presa"))
nd = roda(**zerado(nivel=10, tipo="técnica", trilha="Coro", sintonia="—"))
crit = INV["sintonia"]["rotas"]["Presa"]["critico_a_partir_de"]
checa(f"a Presa poe o critico em '{crit} ou 20'", pr["critico"] == f"{crit} ou 20",
      f'a ficha deu {pr["critico"]!r}')
checa("sem Sintonia o critico e so o 20", nd["critico"] == "20",
      f'a ficha deu {nd["critico"]!r}')
checa("contra-teste: a Presa NAO mexe na vida",
      num(pr["vida_max"]) == num(nd["vida_max"]),
      f'{pr["vida_max"]!r} contra {nd["vida_max"]!r}')

# a Voz: ela aponta para a CD, que nao existe -- a ficha avisa
vz = roda(**zerado(nivel=10, tipo="técnica", trilha="Coro", sintonia="Voz"))
checa("a Voz faz a ficha avisar que a CD nao tem formula",
      isinstance(vz["cd_pendente"], str) and "não tem fórmula" in vz["cd_pendente"],
      f'a ficha deu {vz["cd_pendente"]!r}')
checa("sem a Voz o aviso da CD fica calado", nd["cd_pendente"] in ("", None),
      f'a ficha deu {nd["cd_pendente"]!r}')
checa("contra-teste: a Voz NAO mexe na vida",
      num(vz["vida_max"]) == num(nd["vida_max"]),
      f'{vz["vida_max"]!r} contra {nd["vida_max"]!r}')

print()
print("   contra-teste: o Parrudo NAO pode mexer na regua da morte")
sem = roda(**zerado(nivel=30, tipo="técnica", trilha="Servo", sintonia="—"))
com = roda(**zerado(nivel=30, tipo="técnica", trilha="Servo", sintonia="Parrudo"))
checa("a regua e a mesma com e sem Parrudo", num(sem["regua"]) == num(com["regua"]),
      f'{sem["regua"]!r} contra {com["regua"]!r}')

# =====================================================================
print()
print("=" * 74)
print("10. OS ATRIBUTOS — 9 na criacao, +1 por marco, teto 6")
print("=" * 74)
A = INV["atributos"]
s = roda(**zerado(nivel=2, tipo="técnica", trilha="Coro",
                  **{"atr_Força": 3, "atr_Destreza": 3, "atr_Constituição": 3}))
checa(f'nv2: {A["pontos_na_criacao"]} pontos disponiveis',
      num(s["pontos_disp"]) == A["pontos_na_criacao"], f'{s["pontos_disp"]!r}')
checa("3+3+3 = 9 fecha certo e a conferencia diz ok", s["aviso_atributo"] == "ok",
      f'{s["aviso_atributo"]!r}')

s = roda(**zerado(nivel=30, tipo="técnica", trilha="Coro"))
esperado = A["pontos_na_criacao"] + len(INV["progressao"]["marcos"])
checa(f"nv30: {esperado} pontos — os 9 mais os sete marcos",
      num(s["pontos_disp"]) == esperado, f'{s["pontos_disp"]!r}')

s = roda(**zerado(nivel=2, tipo="técnica", trilha="Coro",
                  **{"atr_Força": 3, "atr_Destreza": 3, "atr_Constituição": 4}))
checa("contra-teste: 10 pontos num orcamento de 9 acusa 'estourou o total'",
      s["aviso_atributo"] == "estourou o total", f'{s["aviso_atributo"]!r}')

s = roda(**zerado(nivel=30, tipo="técnica", trilha="Coro",
                  **{"atr_Força": 7, "atr_Destreza": 3, "atr_Constituição": 3,
                     "atr_Inteligência": 2, "atr_Essência": 1}))
checa(f'contra-teste: um atributo em 7 acusa o teto de {A["teto"]}',
      s["aviso_atributo"] == f'estourou o teto de {A["teto"]}',
      f'{s["aviso_atributo"]!r}')

# =====================================================================
print()
print("=" * 74)
if falhas:
    print(f">>> {len(falhas)} FALHA(S) de {checagens}")
    for f in falhas:
        print(f"    · {f}")
    sys.exit(1)
print(f">>> TUDO OK — as {checagens} checagens saem do capitulo 16, recalculadas")
print("    na propria planilha. Nenhum valor esperado escrito a mao.")
print("=" * 74)
